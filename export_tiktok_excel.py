"""Export every TikTok ad in the DB to an Excel workbook with:
   - one row per ad
   - clickable hyperlinks for the library / profile / video / image URLs
   - colour-coded match_type for fast triage
   - transcript text for each ad that has one"""
import sys, sqlite3, json, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

BASE = os.path.dirname(os.path.abspath(__file__))
DB   = os.path.join(BASE, 'politician_ads.db')
OUT  = os.path.join(BASE, 'tiktok_ads_full.xlsx')

# If the file is open in Excel, fall back to a date-stamped name so we don't crash.
def _open_for_write(path: str) -> str:
    try:
        with open(path, 'ab'):
            return path
    except PermissionError:
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M')
        alt = path.replace('.xlsx', f'_{ts}.xlsx')
        print(f"  [warn] {path} is locked — writing to {alt}")
        return alt
OUT = _open_for_write(OUT)

c = sqlite3.connect(DB)
rows = list(c.execute("""
    SELECT match_type, advertiser_disclosed_name, matched_candidate, matched_party,
           matched_district, ad_id, first_shown, last_shown, ad_status, status_statement,
           reach_raw, times_shown_lower_bound, times_shown_upper_bound,
           ad_funded_by, videos_json, image_urls_json, transcript, transcript_lang_prob,
           ad_url, country_code, checked_at
    FROM tiktok_ads
    ORDER BY
      CASE match_type
        WHEN 'manual_resume' THEN 1
        WHEN 'needs_profile_verification' THEN 2
        WHEN 'rescreen_lastname' THEN 3
        WHEN 'rescreen_lastname_with_initial' THEN 4
        WHEN 'likely_false_positive_content_not_political' THEN 5
        WHEN 'likely_false_positive_first_name_mismatch' THEN 6
        ELSE 9 END,
      advertiser_disclosed_name,
      first_shown
"""))

wb = Workbook()
ws = wb.active
ws.title = "TikTok ads"

# ── Colour palette per match_type ─────────────────────────────────────────────
FILLS = {
    'manual_resume':                              PatternFill('solid', fgColor='C6E0B4'),   # green — confirmed
    'needs_profile_verification':                  PatternFill('solid', fgColor='FFF2CC'),   # yellow — verify
    'rescreen_lastname':                           PatternFill('solid', fgColor='FCE4D6'),   # peach — uncertain
    'rescreen_lastname_with_initial':              PatternFill('solid', fgColor='FCE4D6'),
    'likely_false_positive_content_not_political': PatternFill('solid', fgColor='F4B0B0'),   # red — debunked
    'likely_false_positive_first_name_mismatch':   PatternFill('solid', fgColor='F4B0B0'),
}
HEADER_FILL = PatternFill('solid', fgColor='305496')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
HYPER_FONT  = Font(color='0563C1', underline='single')

HEADERS = [
    'tier', 'match_type', 'handle', 'profile_url',
    'candidate', 'party', 'district',
    'ad_id', 'ad_kind', 'status',
    'first_shown', 'last_shown', 'days',
    'reach_bucket', 'reach_low', 'reach_high',
    'library_url', 'video_cdn_url', 'image_cdn_urls',
    'transcript', 'transcript_lang_prob',
    'ad_funded_by', 'status_statement', 'country', 'checked_at',
]
TIER_LABEL = {
    'manual_resume': '1 — Confirmed',
    'needs_profile_verification': '2 — Verify profile',
    'rescreen_lastname': '3 — Last-name suspect',
    'rescreen_lastname_with_initial': '3 — Last-name suspect',
    'likely_false_positive_content_not_political': '4 — FP (content debunked)',
    'likely_false_positive_first_name_mismatch': '4 — FP (first-name mismatch)',
}

# Write header
for col_idx, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center')

def add_hyperlink(cell, url, display=None):
    if not url: return
    cell.hyperlink = url
    cell.value     = display or url
    cell.font      = HYPER_FONT

# Write data rows
for r_idx, row in enumerate(rows, 2):
    (match_type, handle, cand, party, district, ad_id, first, last, status, stmt,
     reach, lb, ub, funded, vids_j, imgs_j, transcript, tlp,
     ad_url, country, checked_at) = row

    vids = json.loads(vids_j or '[]')
    imgs = json.loads(imgs_j or '[]')
    primary_video = vids[0].get('url') if vids and isinstance(vids[0], dict) else (vids[0] if vids else '')
    kind = 'VIDEO' if vids else ('IMAGE' if imgs else '?')
    profile_url = f"https://www.tiktok.com/@{handle}" if handle else ''
    library_url = f"https://library.tiktok.com/ads/detail/?ad_id={ad_id}"

    try:
        from datetime import date
        days = (date.fromisoformat(last) - date.fromisoformat(first)).days + 1 if first and last else None
    except Exception:
        days = None

    img_concat = '\n'.join(imgs) if imgs else ''

    values = [
        TIER_LABEL.get(match_type, match_type),
        match_type,
        handle or '',
        profile_url,
        cand or '',
        party or '',
        district or '',
        ad_id or '',
        kind,
        status or '',
        first or '',
        last or '',
        days,
        reach or '',
        lb,
        ub,
        library_url,
        primary_video,
        img_concat,
        transcript or '',
        tlp,
        funded or '',
        stmt or '',
        country or '',
        checked_at or '',
    ]

    fill = FILLS.get(match_type)
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=r_idx, column=col_idx, value=v)
        if fill:
            cell.fill = fill

    # Make URL cells clickable
    add_hyperlink(ws.cell(row=r_idx, column=HEADERS.index('profile_url')+1),    profile_url)
    add_hyperlink(ws.cell(row=r_idx, column=HEADERS.index('library_url')+1),    library_url, display='library')
    if primary_video:
        add_hyperlink(ws.cell(row=r_idx, column=HEADERS.index('video_cdn_url')+1), primary_video, display='video')
    if img_concat:
        # show all image URLs separated by newlines, but only hyperlink the first one cleanly
        cell = ws.cell(row=r_idx, column=HEADERS.index('image_cdn_urls')+1)
        cell.hyperlink = imgs[0]
        cell.value = f"image ({len(imgs)} url{'s' if len(imgs)>1 else ''})"
        cell.font = HYPER_FONT

# ── Column widths ─────────────────────────────────────────────────────────────
WIDTHS = {
    'tier': 24, 'match_type': 38, 'handle': 22, 'profile_url': 36,
    'candidate': 22, 'party': 18, 'district': 12,
    'ad_id': 18, 'ad_kind': 7, 'status': 8,
    'first_shown': 11, 'last_shown': 11, 'days': 6,
    'reach_bucket': 12, 'reach_low': 10, 'reach_high': 10,
    'library_url': 14, 'video_cdn_url': 14, 'image_cdn_urls': 14,
    'transcript': 60, 'transcript_lang_prob': 18,
    'ad_funded_by': 22, 'status_statement': 12, 'country': 8, 'checked_at': 22,
}
for i, h in enumerate(HEADERS, 1):
    ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(h, 14)

# Freeze the header
ws.freeze_panes = 'A2'

# Autofilter
ws.auto_filter.ref = ws.dimensions

# Wrap text in transcript column
transcript_col = HEADERS.index('transcript') + 1
for r in range(2, ws.max_row + 1):
    ws.cell(row=r, column=transcript_col).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 45

# ── A summary sheet ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2.append(['match_type', 'advertisers', 'ads', 'description'])
descr = {
    'manual_resume': 'Strict full-name match + transcript-verified political content',
    'needs_profile_verification': 'Silent/music-only ads; surname matches a candidate; verify via profile',
    'rescreen_lastname': 'Last-name match only; still needs verification',
    'rescreen_lastname_with_initial': 'Last-name + initial match (e.g. _ele = Έλενα)',
    'likely_false_positive_content_not_political': 'Transcript revealed non-political content (music, business, etc.)',
    'likely_false_positive_first_name_mismatch': "Handle's first name differs from candidate's first name",
}
for mt, d in descr.items():
    n_adv = c.execute("SELECT COUNT(DISTINCT advertiser_id) FROM tiktok_ads WHERE match_type=?", (mt,)).fetchone()[0]
    n_ads = c.execute("SELECT COUNT(*) FROM tiktok_ads WHERE match_type=?", (mt,)).fetchone()[0]
    ws2.append([mt, n_adv, n_ads, d])

for col, width in [(1, 50), (2, 14), (3, 8), (4, 80)]:
    ws2.column_dimensions[get_column_letter(col)].width = width
for cell in ws2[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
ws2.freeze_panes = 'A2'

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Total rows: {len(rows)}, sheets: {wb.sheetnames}")
