"""
Build candidate_ads_combined.xlsx from the single combined politician_ads.db.
"""

import sys
import sqlite3
import os

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter

from utils import (
    PARTY_PAGE_LABEL,
    flag_page, is_business, is_excluded,
    load_exclusions, load_page_categories, is_non_political_by_category,
)

BASE     = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE, "politician_ads.db")
OUT_FILE = os.path.join(BASE, "candidate_ads_combined.xlsx")

# ── Load exclusions & categories ─────────────────────────────────────────────
EXCLUDE_PAGE_IDS, EXCLUDE_PAGE_NAMES = load_exclusions()
print(f"  Loaded {len(EXCLUDE_PAGE_IDS)} excluded page IDs + {len(EXCLUDE_PAGE_NAMES)} excluded page names")

PAGE_CATEGORIES = load_page_categories()
if PAGE_CATEGORIES:
    print(f"  Loaded {len(PAGE_CATEGORIES)} page categories from cache")

# ── Columns ───────────────────────────────────────────────────────────────────
COLUMNS = [
    "Candidate", "Party", "District",
    "Page Name", "Page ID", "Bylines", "Is Third Party",
    "Start Date", "Stop Date",
    "Impressions Min", "Impressions Max",
    "Spend Min", "Spend Max", "Currency",
    "View Ad", "Checked At", "Source", "Flag", "Removed",
]

# ── Load combined DB ──────────────────────────────────────────────────────────
print("Loading combined DB …")
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
all_rows_raw = conn.execute(
    "SELECT * FROM politician_ads WHERE ad_start_date >= '2025-10-01'"
).fetchall()
conn.close()

def row_to_dict(r):
    d = dict(r)
    src = (d.get("source") or "greek").strip().lower()
    d["_source"] = src.capitalize()
    return d

all_rows = [row_to_dict(r) for r in all_rows_raw]
greek_count = sum(1 for r in all_rows if r["_source"] == "Greek")
latin_count = sum(1 for r in all_rows if r["_source"] == "Latin")
print(f"  {len(all_rows):,} rows  (Greek: {greek_count:,}  Latin: {latin_count:,})")

# ── Apply filters ─────────────────────────────────────────────────────────────
filtered = []
skip_biz = skip_excl = skip_cat = 0
for r in all_rows:
    pid = str(r.get("page_id") or "").strip()
    pn  = r.get("page_name") or ""
    if is_excluded(pid, pn, EXCLUDE_PAGE_IDS, EXCLUDE_PAGE_NAMES):
        skip_excl += 1
        continue
    if is_business(pn):
        skip_biz += 1
        continue
    if is_non_political_by_category(pid, PAGE_CATEGORIES):
        skip_cat += 1
        continue
    filtered.append(r)

print(f"  Excluded (blacklist): {skip_excl}")
print(f"  Excluded (business):  {skip_biz}")
if skip_cat:
    print(f"  Excluded (API cat):   {skip_cat}")
print(f"  Kept:                 {len(filtered):,}")

# ── Build Excel ───────────────────────────────────────────────────────────────
print("Writing Excel …")
wb = Workbook()
ws = wb.active
ws.title = "Combined Ads"

# Header
header_font = Font(bold=True)
for col_idx, col_name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

# Rows
for r in filtered:
    query          = r.get("politician_query", "")
    candidate_name = query.split("|")[0]
    ad_id          = r.get("ad_archive_id") or ""
    public_url     = f"https://www.facebook.com/ads/library/?id={ad_id}" if ad_id else ""
    flag           = flag_page(r.get("page_name"), r.get("politician_query"))

    # Party/political pages get a unified display name
    page_name_display = (
        PARTY_PAGE_LABEL if flag == 'PARTY_PAGE'
        else r.get("page_name")
    )

    removed = r.get("removed")
    removed_label = "YES" if removed == 1 else ("" if removed is None else "")

    row_data = [
        candidate_name,
        r.get("party"),
        r.get("district"),
        page_name_display,
        r.get("page_id"),
        r.get("bylines"),
        r.get("is_third_party"),
        r.get("ad_start_date"),
        r.get("ad_stop_date"),
        r.get("impressions_min"),
        r.get("impressions_max"),
        r.get("spend_min"),
        r.get("spend_max"),
        r.get("currency"),
        "View Ad",
        r.get("checked_at"),
        r.get("_source"),
        flag,
        removed_label,
    ]

    row_idx = ws.max_row + 1
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if col_idx == 15 and public_url:   # "View Ad" column
            cell.hyperlink = public_url
            cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="center")

# Column widths
col_widths = [30, 10, 15, 40, 20, 20, 14, 12, 12, 14, 14, 10, 10, 8, 60, 22, 8, 12, 9]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"

# ── Summary sheet ─────────────────────────────────────────────────────────────
print("Writing Summary sheet …")
ws2 = wb.create_sheet(title="Summary")

SUM_COLUMNS = [
    "Candidate", "Party", "District",
    "Total Ads", "Active Ads", "Inactive Ads", "Removed Ads",
    "Impressions Min (total)", "Impressions Max (total)",
    "Spend Min (total)", "Spend Max (total)", "Currency",
    "Unique Pages", "Latin Ads", "Ad Library Link",
]

# Header
for col_idx, col_name in enumerate(SUM_COLUMNS, 1):
    cell = ws2.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

# Aggregate per candidate
from collections import defaultdict

cand_data: dict = defaultdict(lambda: {
    "party": "", "district": "",
    "total": 0, "active": 0, "inactive": 0, "removed": 0,
    "impr_min": 0, "impr_max": 0,
    "spend_min": 0, "spend_max": 0,
    "currency": "",
    "pages": set(),
    "latin": 0,
})

for r in filtered:
    key  = r.get("politician_query", "")
    name = key.split("|")[0]
    d    = cand_data[name]
    d["party"]    = r.get("party") or d["party"]
    d["district"] = r.get("district") or d["district"]
    d["total"]   += 1
    is_removed  = r.get("removed") == 1
    has_stopped = bool(r.get("ad_stop_date"))
    if is_removed:
        d["removed"]  += 1
    elif has_stopped:
        d["inactive"] += 1
    else:
        d["active"]   += 1
    d["impr_min"]  += r.get("impressions_min") or 0
    d["impr_max"]  += r.get("impressions_max") or 0
    d["spend_min"] += r.get("spend_min") or 0
    d["spend_max"] += r.get("spend_max") or 0
    if r.get("currency"):
        d["currency"] = r["currency"]
    pid = str(r.get("page_id") or "").strip()
    if pid and pid != "0":
        d["pages"].add(pid)
    if (r.get("_source") or "").lower() == "latin":
        d["latin"] += 1

# Sort by total ads descending
sorted_cands = sorted(cand_data.items(), key=lambda x: x[1]["total"], reverse=True)

right = Alignment(horizontal="right")

for name, d in sorted_cands:
    page_id = next(iter(d["pages"]), None)
    ad_lib_url = (
        f"https://www.facebook.com/ads/library/?active_status=all&ad_type=all"
        f"&country=CY&media_type=all&view_all_page_id={page_id}"
    ) if page_id else ""

    row_idx = ws2.max_row + 1
    row_data = [
        name,
        d["party"],
        d["district"] or None,
        d["total"],
        d["active"],
        d["inactive"],
        d["removed"] or None,
        d["impr_min"] or None,
        d["impr_max"] or None,
        d["spend_min"] or None,
        d["spend_max"] or None,
        d["currency"] or None,
        len(d["pages"]),
        d["latin"] or None,
        "Ad Library" if ad_lib_url else None,
    ]
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if col_idx >= 4:
            cell.alignment = right
        if col_idx == 15 and ad_lib_url:
            cell.hyperlink = ad_lib_url
            cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="center")

# Bold header, freeze, column widths
ws2.freeze_panes = "A2"
sum_widths = [30, 10, 15, 11, 11, 11, 11, 22, 22, 18, 18, 9, 13, 10, 14]
for i, w in enumerate(sum_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT_FILE)
print(f"\nDone! Saved to: {OUT_FILE}")
print(f"Total rows: {len(filtered):,}  |  Candidates in summary: {len(cand_data):,}")

# ── Flag breakdown ────────────────────────────────────────────────────────────
flag_counts: dict = {}
for r in filtered:
    fl = flag_page(r.get("page_name"), r.get("politician_query"))
    flag_counts[fl] = flag_counts.get(fl, 0) + 1
print("\nFlag breakdown:")
for fl, cnt in sorted(flag_counts.items()):
    print(f"  {fl:<12} {cnt:,}")
