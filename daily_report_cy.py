"""
daily_report_cy.py — Cyprus daily ad-monitoring report.
────────────────────────────────────────────────────────
Generates two reports each day:

  1. NEWLY REMOVED ADS
     Active election ads that were detected as removed in the last N hours.
     Grouped by politician / page.

  2. RE-ADVERTISERS
     Politicians who had at least one election ad removed AND have since
     posted NEW ads (ad_start_date after their first detected removal).

Output
  • Console summary (always)
  • Excel file  daily_report_YYYY-MM-DD.xlsx  (use --no-excel to skip)

Usage
    python daily_report_cy.py                  # last 24h, writes Excel
    python daily_report_cy.py --hours 48       # look back 48h
    python daily_report_cy.py --no-excel       # console only
    python daily_report_cy.py --out mydir      # custom output folder
"""

import os, sys, sqlite3, json, argparse
from datetime import datetime, timezone, timedelta

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BASE    = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE, "politician_ads.db")
BL_FILE = os.path.join(BASE, "page_blocklist.json")

AD_LIB_URL = "https://www.facebook.com/ads/library/?id={}"


# ── helpers ───────────────────────────────────────────────────────────────────

def load_blocklist() -> set:
    if os.path.exists(BL_FILE):
        with open(BL_FILE, encoding="utf-8") as f:
            data = json.load(f)
        pages = data.get("pages", data) if isinstance(data, dict) else {}
        return set(str(k) for k in pages)
    return set()


def fmt_name(politician_query: str) -> str:
    """Return 'Name (Party · District)' from 'Name|Party|District'."""
    parts = (politician_query or "").split("|")
    name = parts[0].strip() if parts else politician_query
    party = parts[1].strip() if len(parts) > 1 else ""
    district = parts[2].strip() if len(parts) > 2 else ""
    extra = " · ".join(p for p in [party, district] if p)
    return f"{name}  ({extra})" if extra else name


# ── queries ───────────────────────────────────────────────────────────────────

def get_newly_removed(conn, blocklist: set, hours: int) -> list[dict]:
    """
    Ads detected as removed within the last `hours` hours.
    Excludes blocklisted pages and ads confirmed as non-election-related.
    If election_related column exists, excludes 'NO' rows; includes YES + NULL.
    """
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours)).isoformat()

    # Check if election_related column exists
    cols_in_db = {r[1] for r in conn.execute("PRAGMA table_info(politician_ads)").fetchall()}
    er_filter  = "AND (election_related IS NULL OR election_related != 'NO')" \
                 if "election_related" in cols_in_db else ""

    sql = f"""
        SELECT
            ad_archive_id,
            page_id,
            page_name,
            politician_query,
            party,
            district,
            ad_start_date,
            ad_stop_date,
            removed_checked_at,
            impressions_min,
            impressions_max,
            spend_min,
            spend_max,
            currency
        FROM politician_ads
        WHERE removed = 1
          AND removed_checked_at >= ?
          {er_filter}
        ORDER BY removed_checked_at DESC
    """
    rows = conn.execute(sql, (cutoff,)).fetchall()
    cols = ["ad_archive_id","page_id","page_name","politician_query","party","district",
            "ad_start_date","ad_stop_date","removed_checked_at",
            "impressions_min","impressions_max","spend_min","spend_max","currency"]
    result = []
    for r in rows:
        d = dict(zip(cols, r))
        if str(d["page_id"] or "") not in blocklist:
            result.append(d)
    return result


def get_new_ads(conn, blocklist: set, hours: int) -> list[dict]:
    """
    Ads first seen within the last `hours` hours.
    Uses first_seen_at (set once on INSERT) so the same ad never appears twice.
    Excludes removed ads, NO-classified ads, and blocklisted pages.
    """
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours)).isoformat()

    cols_in_db = {r[1] for r in conn.execute("PRAGMA table_info(politician_ads)").fetchall()}
    er_filter  = "AND (election_related IS NULL OR election_related != 'NO')" \
                 if "election_related" in cols_in_db else ""

    # Use COALESCE so ads with NULL first_seen_at fall back to checked_at.
    # (first_seen_at is only populated for ads inserted after the column was added;
    #  checked_at is always set and updated on every upsert.)
    fsa_expr = "COALESCE(first_seen_at, checked_at)" \
               if "first_seen_at" in cols_in_db else "checked_at"

    sql = f"""
        SELECT
            ad_archive_id,
            page_id,
            page_name,
            politician_query,
            party,
            district,
            ad_start_date,
            ad_stop_date,
            impressions_min,
            impressions_max,
            spend_min,
            spend_max,
            currency,
            {fsa_expr} AS first_seen_at
        FROM politician_ads
        WHERE {fsa_expr} >= ?
          AND removed = 0
          {er_filter}
        ORDER BY {fsa_expr} DESC
    """
    rows = conn.execute(sql, (cutoff,)).fetchall()
    col_names = ["ad_archive_id","page_id","page_name","politician_query","party","district",
                 "ad_start_date","ad_stop_date",
                 "impressions_min","impressions_max","spend_min","spend_max",
                 "currency","first_seen_at"]
    result = []
    for r in rows:
        d = dict(zip(col_names, r))
        if str(d["page_id"] or "") not in blocklist:
            result.append(d)
    return result


def get_readvertisers(conn, blocklist: set) -> list[dict]:
    """
    Pages that:
      (a) have at least one removed election ad  AND
      (b) have at least one active ad whose ad_start_date is AFTER the
          date their first removal was detected.

    Returns the NEW active ads for those pages, annotated with removal info.
    """
    # Check if election_related column exists
    cols_in_db = {r[1] for r in conn.execute("PRAGMA table_info(politician_ads)").fetchall()}
    er_filter_removed = "AND (pa2.election_related IS NULL OR pa2.election_related != 'NO')" \
                        if "election_related" in cols_in_db else ""
    er_filter_new     = "AND (pa.election_related IS NULL OR pa.election_related != 'NO')" \
                        if "election_related" in cols_in_db else ""

    sql = f"""
        WITH first_removal AS (
            -- When did each page first get an ad removed?
            SELECT
                pa2.page_id,
                MIN(pa2.removed_checked_at) AS first_removed_at,
                COUNT(*)                    AS total_removed
            FROM politician_ads pa2
            WHERE pa2.removed = 1
              {er_filter_removed}
            GROUP BY pa2.page_id
        ),
        new_ads AS (
            -- Active ads that started AFTER the page's first removal
            SELECT
                pa.ad_archive_id,
                pa.page_id,
                pa.page_name,
                pa.politician_query,
                pa.party,
                pa.district,
                pa.ad_start_date,
                pa.ad_stop_date,
                pa.impressions_min,
                pa.impressions_max,
                pa.spend_min,
                pa.spend_max,
                pa.currency,
                pa.checked_at,
                fr.first_removed_at,
                fr.total_removed
            FROM politician_ads pa
            JOIN first_removal fr ON pa.page_id = fr.page_id
            WHERE pa.removed = 0
              -- Use first_seen_at (reliable) if available, fall back to ad_start_date
              AND COALESCE(pa.first_seen_at, pa.ad_start_date) > fr.first_removed_at
              {er_filter_new}
        )
        SELECT * FROM new_ads
        ORDER BY page_id, ad_start_date DESC
    """
    rows = conn.execute(sql).fetchall()
    cols = ["ad_archive_id","page_id","page_name","politician_query","party","district",
            "ad_start_date","ad_stop_date",
            "impressions_min","impressions_max","spend_min","spend_max","currency",
            "checked_at","first_removed_at","total_removed"]
    result = []
    for r in rows:
        d = dict(zip(cols, r))
        if str(d["page_id"] or "") not in blocklist:
            result.append(d)
    return result


# ── console output ─────────────────────────────────────────────────────────────

def print_removed(rows: list[dict], hours: int):
    print(f"\n{'═'*70}")
    print(f"  📛  NEWLY REMOVED ADS  (last {hours}h)")
    print(f"{'═'*70}")
    if not rows:
        print("  ✅ No ads detected as removed in this period.\n")
        return

    # Group by politician_query
    from collections import defaultdict
    by_pol = defaultdict(list)
    for r in rows:
        by_pol[r["politician_query"]].append(r)

    for query, ads in sorted(by_pol.items()):
        print(f"\n  👤  {fmt_name(query)}")
        print(f"      Page: {ads[0]['page_name']}  (id {ads[0]['page_id']})")
        for a in ads:
            imp   = f"{a['impressions_min']:,}–{a['impressions_max']:,}" if a['impressions_min'] else "n/a"
            spend = (f"{a['spend_min']}–{a['spend_max']} {a['currency']}"
                     if a['spend_min'] is not None else "n/a")
            chk   = (a['removed_checked_at'] or "")[:16].replace("T", " ")
            print(f"      ❌  {a['ad_archive_id']}  |  started {a['ad_start_date']}  "
                  f"|  imp {imp}  |  spend {spend}  |  detected {chk}")

    print(f"\n  Total: {len(rows)} ad(s) removed across {len(by_pol)} politician(s).\n")


def print_readvertisers(rows: list[dict]):
    print(f"{'═'*70}")
    print(f"  🔄  RE-ADVERTISERS  (new ads after a removal)")
    print(f"{'═'*70}")
    if not rows:
        print("  ✅ No re-advertisers detected.\n")
        return

    from collections import defaultdict
    by_page = defaultdict(list)
    for r in rows:
        by_page[r["page_id"]].append(r)

    for page_id, ads in sorted(by_page.items()):
        sample = ads[0]
        first_rem = (sample["first_removed_at"] or "")[:10]
        print(f"\n  👤  {fmt_name(sample['politician_query'])}")
        print(f"      Page: {sample['page_name']}  (id {page_id})")
        print(f"      ⚠️   First removal detected: {first_rem}  "
              f"|  total removed: {sample['total_removed']}")
        print(f"      📢  New active ads ({len(ads)}):")
        for a in ads[:5]:   # cap display at 5 per page
            imp   = f"{a['impressions_min']:,}–{a['impressions_max']:,}" if a['impressions_min'] else "n/a"
            spend = (f"{a['spend_min']}–{a['spend_max']} {a['currency']}"
                     if a['spend_min'] is not None else "n/a")
            print(f"        ▶  {a['ad_archive_id']}  |  started {a['ad_start_date']}  "
                  f"|  imp {imp}  |  spend {spend}")
        if len(ads) > 5:
            print(f"        … and {len(ads)-5} more (see Excel)")

    from collections import Counter
    pages = Counter(r["page_id"] for r in rows)
    print(f"\n  Total: {len(rows)} new ad(s) from {len(pages)} politician(s) "
          f"who had previous removals.\n")


# ── Excel output ───────────────────────────────────────────────────────────────

SHEET_REMOVED = "Removed Today"
SHEET_READV   = "Re-Advertisers"
SHEET_NEW     = "New Ads Today"

HDR1 = ["Politician", "Party", "District", "Page Name", "Page ID",
        "Ad ID", "View Ad", "Ad Start", "Ad Stop",
        "Impressions", "Spend", "Currency", "Detected At"]
HDR2 = ["Politician", "Party", "District", "Page Name", "Page ID",
        "New Ad ID", "View Ad", "New Ad Start", "New Ad Stop",
        "Impressions", "Spend", "Currency",
        "First Removal Detected", "Total Ads Removed"]
HDR3 = ["Politician", "Party", "District", "Page Name", "Page ID",
        "Ad ID", "View Ad", "Ad Start", "Ad Stop",
        "Impressions", "Spend", "Currency", "First Seen At"]


def write_excel(removed_rows: list[dict], readv_rows: list[dict],
                new_ads_rows: list[dict], hours: int, out_dir: str) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [!] openpyxl not installed — skipping Excel output.")
        return ""

    today = datetime.now().strftime("%Y-%m-%d")
    fname = os.path.join(out_dir, f"daily_report_{today}.xlsx")

    # ── colour palette ──────────────────────────────────────────────────────
    HDR_FILL_RED   = PatternFill("solid", fgColor="C00000")
    HDR_FILL_BLUE  = PatternFill("solid", fgColor="1F4E79")
    HDR_FILL_GREEN = PatternFill("solid", fgColor="375623")
    HDR_FONT      = Font(bold=True, color="FFFFFF", size=10)
    LINK_FONT     = Font(color="0563C1", underline="single")
    ALT_FILL      = PatternFill("solid", fgColor="FFF2CC")

    def set_hdr(ws, row_vals, fill):
        ws.append(row_vals)
        for cell in ws[ws.max_row]:
            cell.font      = HDR_FONT
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.row_dimensions[ws.max_row].height = 22
        ws.freeze_panes = f"A{ws.max_row + 1}"

    def autofit(ws):
        for col in ws.columns:
            length = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(length + 2, 50)

    # ── Load existing file or create fresh ──────────────────────────────────
    if os.path.exists(fname):
        wb = openpyxl.load_workbook(fname)

        # ── Removed sheet: append only new rows ────────────────────────────
        if SHEET_REMOVED in wb.sheetnames:
            ws1 = wb[SHEET_REMOVED]
            # Collect ad IDs already written (column 6 = Ad ID)
            existing_ids = {
                ws1.cell(row=r, column=6).value
                for r in range(2, ws1.max_row + 1)
            }
            # Remove placeholder "no data" row if it's the only content row
            if ws1.max_row == 2 and ws1.cell(row=2, column=6).value is None:
                ws1.delete_rows(2)
        else:
            # Sheet missing in old-format file — create it
            ws1 = wb.create_sheet(SHEET_REMOVED, 0)
            set_hdr(ws1, HDR1, HDR_FILL_RED)
            existing_ids = set()

        new_removed = [r for r in removed_rows
                       if r["ad_archive_id"] not in existing_ids]

        # ── Re-Advertisers sheet: always full refresh ───────────────────────
        if SHEET_READV in wb.sheetnames:
            del wb[SHEET_READV]
        ws2 = wb.create_sheet(SHEET_READV)

        # ── New Ads sheet: append only new rows ─────────────────────────────
        if SHEET_NEW in wb.sheetnames:
            ws3 = wb[SHEET_NEW]
            existing_new_ids = {
                ws3.cell(row=r, column=6).value
                for r in range(2, ws3.max_row + 1)
            }
            if ws3.max_row == 2 and ws3.cell(row=2, column=6).value is None:
                ws3.delete_rows(2)
        else:
            ws3 = wb.create_sheet(SHEET_NEW)
            set_hdr(ws3, HDR3, HDR_FILL_GREEN)
            existing_new_ids = set()
        new_new_ads = [r for r in new_ads_rows
                       if r["ad_archive_id"] not in existing_new_ids]

    else:
        # Brand-new file for today
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = SHEET_REMOVED
        set_hdr(ws1, HDR1, HDR_FILL_RED)
        new_removed = removed_rows
        ws2 = wb.create_sheet(SHEET_READV)
        ws3 = wb.create_sheet(SHEET_NEW)
        set_hdr(ws3, HDR3, HDR_FILL_GREEN)
        new_new_ads = new_ads_rows

    # ── Append new removed rows ──────────────────────────────────────────────
    for r in new_removed:
        row_num = ws1.max_row + 1
        name  = (r["politician_query"] or "").split("|")[0].strip()
        imp   = (f"{r['impressions_min']:,}–{r['impressions_max']:,}"
                 if r['impressions_min'] else "")
        spend = (f"{r['spend_min']}–{r['spend_max']}"
                 if r['spend_min'] is not None else "")
        chk   = (r['removed_checked_at'] or "")[:16].replace("T", " ")
        ws1.append([name, r["party"], r["district"], r["page_name"], r["page_id"],
                    r["ad_archive_id"], "View", r["ad_start_date"], r["ad_stop_date"],
                    imp, spend, r["currency"], chk])
        c = ws1.cell(row=row_num, column=7)
        c.hyperlink = AD_LIB_URL.format(r["ad_archive_id"])
        c.font      = LINK_FONT
        if row_num % 2 == 0:
            for col in range(1, len(HDR1) + 1):
                ws1.cell(row=row_num, column=col).fill = ALT_FILL

    # Placeholder only if the sheet is still completely empty (header only)
    if ws1.max_row == 1:
        ws1.append(["No ads detected as removed today."])

    autofit(ws1)
    print(f"  [excel] Removed sheet: +{len(new_removed)} new row(s)  "
          f"(total rows: {ws1.max_row - 1})")

    # ── Re-Advertisers sheet (full refresh each run) ─────────────────────────
    set_hdr(ws2, HDR2, HDR_FILL_BLUE)

    for i, r in enumerate(readv_rows, 2):
        name  = (r["politician_query"] or "").split("|")[0].strip()
        imp   = (f"{r['impressions_min']:,}–{r['impressions_max']:,}"
                 if r['impressions_min'] else "")
        spend = (f"{r['spend_min']}–{r['spend_max']}"
                 if r['spend_min'] is not None else "")
        first = (r["first_removed_at"] or "")[:10]
        ws2.append([name, r["party"], r["district"], r["page_name"], r["page_id"],
                    r["ad_archive_id"], "View", r["ad_start_date"], r["ad_stop_date"],
                    imp, spend, r["currency"], first, r["total_removed"]])
        c = ws2.cell(row=i, column=7)
        c.hyperlink = AD_LIB_URL.format(r["ad_archive_id"])
        c.font      = LINK_FONT
        if i % 2 == 0:
            for col in range(1, len(HDR2) + 1):
                ws2.cell(row=i, column=col).fill = ALT_FILL

    if not readv_rows:
        ws2.append(["No re-advertisers detected."])

    autofit(ws2)

    # ── Append new ads rows (sheet 3) ────────────────────────────────────────
    for r in new_new_ads:
        row_num = ws3.max_row + 1
        name  = (r["politician_query"] or "").split("|")[0].strip()
        imp   = (f"{r['impressions_min']:,}–{r['impressions_max']:,}"
                 if r['impressions_min'] else "")
        spend = (f"{r['spend_min']}–{r['spend_max']}"
                 if r['spend_min'] is not None else "")
        seen  = (r["first_seen_at"] or "")[:16].replace("T", " ")
        ws3.append([name, r["party"], r["district"], r["page_name"], r["page_id"],
                    r["ad_archive_id"], "View", r["ad_start_date"], r["ad_stop_date"],
                    imp, spend, r["currency"], seen])
        c = ws3.cell(row=row_num, column=7)
        c.hyperlink = AD_LIB_URL.format(r["ad_archive_id"])
        c.font      = LINK_FONT
        if row_num % 2 == 0:
            for col in range(1, len(HDR3) + 1):
                ws3.cell(row=row_num, column=col).fill = ALT_FILL

    if ws3.max_row == 1:
        ws3.append(["No new ads detected today."])

    autofit(ws3)
    print(f"  [excel] New Ads sheet:    +{len(new_new_ads)} new row(s)  "
          f"(total rows: {ws3.max_row - 1})")

    wb.save(fname)
    return fname


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Cyprus daily ad monitoring report.")
    parser.add_argument("--hours",    type=int,  default=24,
                        help="Look-back window for newly removed ads (default 24)")
    parser.add_argument("--no-excel", action="store_true",
                        help="Skip Excel output, print to console only")
    parser.add_argument("--out",      default=BASE,
                        help="Output directory for Excel (default: script folder)")
    args = parser.parse_args()

    print(f"\n{'═'*70}")
    print(f"  Cyprus Daily Ad Report — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"{'═'*70}")

    blocklist = load_blocklist()
    conn      = sqlite3.connect(DB_PATH)

    removed_rows  = get_newly_removed(conn, blocklist, args.hours)
    readv_rows    = get_readvertisers(conn, blocklist)
    new_ads_rows  = get_new_ads(conn, blocklist, args.hours)

    conn.close()

    print_removed(removed_rows, args.hours)
    print_readvertisers(readv_rows)

    if not args.no_excel:
        path = write_excel(removed_rows, readv_rows, new_ads_rows, args.hours, args.out)
        if path:
            print(f"  💾  Excel saved → {path}\n")

    print(f"{'═'*70}")
    print(f"  Summary:")
    print(f"    Newly removed (last {args.hours}h) : {len(removed_rows)}")
    print(f"    New ads (last {args.hours}h)        : {len(new_ads_rows)}")
    print(f"    Re-advertiser new ads           : {len(readv_rows)}")
    print(f"{'═'*70}\n")


if __name__ == "__main__":
    main()
