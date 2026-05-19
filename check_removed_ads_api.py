"""
check_removed_ads_api.py — API-based removal detection for CY + MT.

Replaces the Playwright-based check_removed_ads_cy.py / check_removed_ads_mt.py.

Meta's Ads Library API returns the removal notice directly inside ad_creative_bodies
for ads that were taken down for violating Advertising Standards — no browser needed.
This eliminates bot-detection failures on GitHub Actions.

Strategy:
  1. Load unchecked / stale-active ads from both DBs.
  2. Group by page_id.
  3. For each page: query the API with ad_creative_bodies.
  4. Detect removal; update DB (never downgrade removed=1→0).

Usage:
    python check_removed_ads_api.py                      # CY + MT, unchecked + stale
    python check_removed_ads_api.py --cy                 # CY only
    python check_removed_ads_api.py --mt                 # MT only
    python check_removed_ads_api.py --all                # re-check all active ads
    python check_removed_ads_api.py --active-only        # fast mode: only pages with running ads
    python check_removed_ads_api.py --limit 1000         # cap total ads checked
    python check_removed_ads_api.py --recheck-days 7
    python check_removed_ads_api.py --sleep 2            # seconds between page requests

--active-only mode:
    Restricts checking to pages that currently have at least one running ad
    (ad_stop_date IS NULL or >= today).  Only checks non-removed ads on those pages.
    CY: ~26 pages (~78 sec).  MT: ~30 pages (~90 sec).  Safe to run hourly.
"""

import os, sys, sqlite3, json, time, argparse, requests, contextlib
from collections import defaultdict
from datetime import datetime, timezone, timedelta, date
from dotenv import load_dotenv

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

load_dotenv(override=True)

BASE            = os.path.dirname(os.path.abspath(__file__))
CY_DB           = os.path.join(BASE, "politician_ads.db")
MT_DB           = os.path.join(BASE, "politician_ads_mt.db")
META_URL               = "https://graph.facebook.com/v25.0/ads_archive"
LOG_FILE               = os.path.join(BASE, "removals_log.xlsx")
# Dedicated state files for page cursors (--max-pages mode).
# Kept separate from fetch_state.json / fetch_state_mt.json so that
# check_removed (group: db-write) never races with cy_refresh / mt_refresh
# (groups: cy-db-write / mt-db-write) over the same JSON file.
CHECK_REMOVED_STATE_CY = os.path.join(BASE, "check_removed_state.json")
CHECK_REMOVED_STATE_MT = os.path.join(BASE, "check_removed_state_mt.json")
AD_URL    = "https://www.facebook.com/ads/library/?id={}"

LOG_HEADERS = [
    "Date Detected (UTC)", "Country", "Candidate", "Party", "District",
    "Page Name", "Ad ID", "Ad Start Date", "Ad End Date",
    "Spend Min (€)", "Spend Max (€)", "Impressions Max", "Ad Library URL",
]

REMOVAL_TEXT = "this content was removed because it didn't follow our advertising standards"


# ── DB ────────────────────────────────────────────────────────────────────────

def load_active_page_ads(conn) -> list[dict]:
    """
    Fast-check mode (--active-only): return all non-removed ads that belong to
    pages currently running at least one ad (stop_date IS NULL or >= today).

    This restricts the API call queue to only the pages that matter right now,
    making the check fast enough to run hourly without hitting rate limits.
    """
    today_str = str(date.today())
    rows = conn.execute("""
        SELECT a.ad_archive_id, a.page_id, a.ad_start_date
        FROM politician_ads a
        WHERE a.page_id IN (
            SELECT DISTINCT page_id
            FROM politician_ads
            WHERE (ad_stop_date IS NULL OR ad_stop_date = '' OR ad_stop_date >= ?)
              AND election_related IN ('YES', 'UNCERTAIN')
              AND page_id IS NOT NULL AND page_id != ''
        )
        AND a.election_related IN ('YES', 'UNCERTAIN')
        AND (a.removed IS NULL OR a.removed = 0)
        ORDER BY a.ad_start_date DESC
    """, (today_str,)).fetchall()
    return [{'ad_archive_id': r[0], 'page_id': str(r[1] or ''), 'ad_start_date': r[2] or ''}
            for r in rows]


def load_unchecked(conn, only_unchecked: bool, recheck_days: float,
                   since: str, limit: int) -> list[dict]:
    """
    Return ads that need a removal check.

    only_unchecked=True  → never-checked + stale active (removed=0 checked > recheck_days ago)
    only_unchecked=False → all active ads (--all mode)
    """
    today_str = str(date.today())

    er_filter = "AND election_related IN ('YES', 'UNCERTAIN')"

    if only_unchecked:
        if recheck_days > 0:
            cutoff = (
                datetime.now(timezone.utc) - timedelta(days=recheck_days)
            ).isoformat()
            time_clause = f"""(
                removed_checked_at IS NULL
                OR (
                    removed = 0
                    AND removed_checked_at < '{cutoff}'
                    AND (ad_stop_date IS NULL OR ad_stop_date = '' OR ad_stop_date >= '{today_str}')
                )
            )"""
        else:
            time_clause = "removed_checked_at IS NULL"

        sql = f"""
            SELECT ad_archive_id, page_id, ad_start_date
            FROM politician_ads
            WHERE {time_clause}
              {er_filter}
            ORDER BY ad_start_date DESC
        """
    else:
        # --all: re-check everything that's still active
        sql = f"""
            SELECT ad_archive_id, page_id, ad_start_date
            FROM politician_ads
            WHERE (ad_stop_date IS NULL OR ad_stop_date = '' OR ad_stop_date >= '{today_str}')
              {er_filter}
            ORDER BY ad_start_date DESC
        """

    rows = conn.execute(sql).fetchall()
    ads = [{'ad_archive_id': r[0], 'page_id': str(r[1] or ''), 'ad_start_date': r[2] or ''}
           for r in rows]

    if since:
        ads = [a for a in ads if a['ad_start_date'] >= since]

    if limit:
        ads = ads[:limit]

    return ads


# ── Page-cursor helpers (for --max-pages mode) ───────────────────────────────

def load_page_cursor(state_file: str, key: str) -> str:
    """Return the page_id of the first page to process next run.
    Empty string means start from the beginning of the sorted list."""
    if os.path.exists(state_file):
        try:
            with open(state_file, encoding='utf-8') as f:
                return str(json.load(f).get(key, ""))
        except Exception:
            pass
    return ""


def save_page_cursor(state_file: str, key: str, value: str) -> None:
    """Persist the page cursor for the next run (atomic write)."""
    state = {}
    if os.path.exists(state_file):
        try:
            with open(state_file, encoding='utf-8') as f:
                state = json.load(f)
        except Exception:
            pass
    state[key] = value
    tmp = state_file + ".tmp"
    try:
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2)
        os.replace(tmp, state_file)
    except Exception as e:
        print(f"  [cursor] WARNING: could not save cursor: {e}")


def save_results(conn, results: list[tuple]) -> None:
    """
    Bulk-save [(ad_archive_id, removed_int, now_str, spend_min, spend_max,
                imp_min, imp_max, currency), ...].
    Policy: removed=1 is never downgraded to 0.
    Spend/impressions are updated whenever the API returns them (fills NULLs).
    """
    for row in results:
        ad_id, removed, ts = row[0], row[1], row[2]
        smin, smax = row[3], row[4]
        imin, imax = row[5], row[6]
        cur        = row[7] if len(row) > 7 else None

        if removed == 1:
            conn.execute("""
                UPDATE politician_ads
                SET removed=1, removed_checked_at=?,
                    spend_min      = CASE WHEN ? IS NOT NULL THEN ? ELSE spend_min END,
                    spend_max      = CASE WHEN ? IS NOT NULL THEN ? ELSE spend_max END,
                    impressions_min= CASE WHEN ? IS NOT NULL THEN ? ELSE impressions_min END,
                    impressions_max= CASE WHEN ? IS NOT NULL THEN ? ELSE impressions_max END,
                    currency       = CASE WHEN ? IS NOT NULL THEN ? ELSE currency END
                WHERE ad_archive_id=?
            """, (ts,
                  smin, smin, smax, smax,
                  imin, imin, imax, imax,
                  cur,  cur,
                  ad_id))
        else:
            conn.execute("""
                UPDATE politician_ads
                SET removed=0, removed_checked_at=?,
                    spend_min      = CASE WHEN ? IS NOT NULL THEN ? ELSE spend_min END,
                    spend_max      = CASE WHEN ? IS NOT NULL THEN ? ELSE spend_max END,
                    impressions_min= CASE WHEN ? IS NOT NULL THEN ? ELSE impressions_min END,
                    impressions_max= CASE WHEN ? IS NOT NULL THEN ? ELSE impressions_max END,
                    currency       = CASE WHEN ? IS NOT NULL THEN ? ELSE currency END
                WHERE ad_archive_id=? AND (removed IS NULL OR removed = 0)
            """, (ts,
                  smin, smin, smax, smax,
                  imin, imin, imax, imax,
                  cur,  cur,
                  ad_id))
    conn.commit()


# ── Excel log ────────────────────────────────────────────────────────────────

@contextlib.contextmanager
def _log_file_lock():
    """
    Cross-process exclusive lock for removals_log.xlsx.

    Both check_removed.yml (db-write group) and check_removed_active.yml
    (active-check group) can detect removals simultaneously.  Without this
    lock, the second process's openpyxl .save() overwrites the first one's
    newly appended rows — silently losing removal log entries.

    Uses fcntl.LOCK_EX on Linux/macOS (GitHub Actions = ubuntu-latest).
    Falls back to a no-op on Windows so local development still works.
    """
    lockfile = LOG_FILE + ".lock"
    fd = open(lockfile, "w")
    try:
        try:
            import fcntl
            fcntl.flock(fd, fcntl.LOCK_EX)   # blocks until acquired
        except ImportError:
            pass                              # Windows: accept the race
        yield
    finally:
        try:
            import fcntl
            fcntl.flock(fd, fcntl.LOCK_UN)
        except ImportError:
            pass
        fd.close()


def _load_log_ids() -> set:
    """Return ad_archive_ids already in removals_log.xlsx (avoid duplicate rows)."""
    if not os.path.exists(LOG_FILE):
        return set()
    try:
        from openpyxl import load_workbook
        wb = load_workbook(LOG_FILE, read_only=True)
        ws = wb.active
        ids = set()
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue   # skip header
            ad_id = row[6]  # column G = Ad ID
            if ad_id:
                ids.add(str(ad_id))
        wb.close()
        return ids
    except Exception as e:
        print(f"  [log] Could not read existing log: {e}")
        return set()


def _append_to_log(rows: list[dict]) -> None:
    """
    Append new-removal rows to removals_log.xlsx, creating it if needed.

    Holds an exclusive file lock for the entire read-modify-write cycle so
    that concurrent runs of check_removed and check_removed_active cannot
    overwrite each other's rows.  IDs are re-checked under the lock to
    prevent duplicate entries when both processes detect the same removal.
    """
    if not rows:
        return
    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        RED_FILL  = PatternFill("solid", fgColor="C00000")
        WHITE_HDR = Font(bold=True, color="FFFFFF")

        with _log_file_lock():
            # Re-read existing IDs under the lock: if another process already
            # wrote some of these IDs between our _load_log_ids() call and now,
            # skip them to avoid duplicate rows.
            existing_ids_now = _load_log_ids()
            rows = [r for r in rows if str(r["ad_id"]) not in existing_ids_now]
            if not rows:
                return

            if os.path.exists(LOG_FILE):
                wb = load_workbook(LOG_FILE)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Removals Log"
                ws.append(LOG_HEADERS)
                for cell in ws[1]:
                    cell.fill      = RED_FILL
                    cell.font      = WHITE_HDR
                    cell.alignment = Alignment(horizontal="center", wrap_text=True)
                ws.row_dimensions[1].height = 22
                ws.freeze_panes = "A2"
                widths = [20, 8, 28, 16, 14, 32, 20, 12, 12, 12, 12, 14, 55]
                for i, w in enumerate(widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = w

            for r in rows:
                ws.append([
                    r["detected_at"], r["country"], r["candidate"],
                    r["party"], r["district"], r["page_name"],
                    r["ad_id"], r["start_date"], r["stop_date"],
                    r["spend_min"], r["spend_max"], r["impressions_max"],
                    AD_URL.format(r["ad_id"]),
                ])
                ws.cell(ws.max_row, 13).font = Font(color="0563C1", underline="single")

            wb.save(LOG_FILE)
            print(f"  [log] Appended {len(rows)} new removal(s) → removals_log.xlsx")

    except ImportError:
        print("  [log] openpyxl not installed — skipping Excel log")
    except Exception as e:
        print(f"  [log] Error writing log: {e}")


def log_new_removals(conn, newly_removed_ids: list[str], country: str,
                     detected_at: str, existing_log_ids: set) -> None:
    """Look up details for newly removed ads and append to the log."""
    to_log = [aid for aid in newly_removed_ids if aid not in existing_log_ids]
    if not to_log:
        return

    placeholders = ",".join("?" * len(to_log))
    rows_db = conn.execute(f"""
        SELECT ad_archive_id, politician_query, party, page_name,
               ad_start_date, ad_stop_date,
               spend_min, spend_max, impressions_max
        FROM politician_ads
        WHERE ad_archive_id IN ({placeholders})
    """, to_log).fetchall()

    log_rows = []
    for r in rows_db:
        query  = r[1] or ""
        parts  = query.split("|")
        cand   = parts[0].strip()
        party  = r[2] or (parts[1].strip() if len(parts) > 1 else "")
        dist   = parts[2].strip() if len(parts) > 2 else ""
        log_rows.append({
            "detected_at":   detected_at[:19].replace("T", " "),
            "country":       country,
            "candidate":     cand,
            "party":         party,
            "district":      dist,
            "page_name":     r[3] or "",
            "ad_id":         r[0],
            "start_date":    r[4] or "",
            "stop_date":     r[5] or "",
            "spend_min":     r[6],
            "spend_max":     r[7],
            "impressions_max": r[8],
        })

    _append_to_log(log_rows)
    # Update the in-memory set so subsequent DBs don't re-log the same IDs
    existing_log_ids.update(r["ad_id"] for r in log_rows)


# ── API ───────────────────────────────────────────────────────────────────────

RATE_LIMIT_CODE      = 613   # Meta API error code for rate limiting
RATE_LIMIT_SLEEP     = 30   # seconds to wait on rate limit before retry
MAX_RETRIES          = 1    # one retry then skip — keeps CI job within timeout
CONSEC_ERROR_ABORT   = 5    # abort this DB after N consecutive non-rate-limit failures
SERVER_ERROR_SLEEP   = 10   # seconds to wait before retrying a 500


def fetch_page_ads(page_id: str, since_date: str, country: str, token: str):
    """
    Query the Ads Library API for all ads on page_id since since_date.

    Returns (results, seen_ids, had_error):
      results  – {ad_archive_id: {is_removed, spend_min, ...}} for ads WITH bodies
      seen_ids – set of ALL ad IDs the API returned (with or without bodies)
      had_error – True if the API call failed (caller should skip, not conclude removal)

    Removal is detected two ways:
      1. Body contains REMOVAL_TEXT  → is_removed=True in results
      2. Ad ID in seen_ids but has no body  → treated as removed (content deleted by Meta)
      3. Ad ID absent from seen_ids entirely, but seen_ids is non-empty (page has other ads)
         → caller treats this as removed too
    """
    params = {
        "ad_type":              "ALL",
        "ad_reached_countries": json.dumps([country]),
        "search_page_ids":      json.dumps([page_id]),
        "ad_delivery_date_min": since_date or "2025-01-01",
        "fields":               "id,ad_creative_bodies,spend,impressions,currency",
        "limit":                100,
        "access_token":         token,
    }
    results  = {}
    seen_ids = set()   # every ad ID the API returned, regardless of body
    url      = META_URL
    retries  = 0
    try:
        while url:
            resp = requests.get(url, params=params, timeout=30)
            if resp.status_code != 200:
                try:
                    err_body = resp.json().get("error", {})
                except Exception:
                    err_body = {}
                err_code = err_body.get("code", 0)
                err_msg  = err_body.get("message", f"HTTP {resp.status_code}")
                if err_code == RATE_LIMIT_CODE and retries < MAX_RETRIES:
                    wait = RATE_LIMIT_SLEEP * (2 ** retries)
                    print(f"      [rate limit] sleeping {wait}s then retrying page {page_id}…")
                    time.sleep(wait)
                    retries += 1
                    continue
                if resp.status_code >= 500 and retries < 1:
                    print(f"      [!] API {resp.status_code} — sleeping {SERVER_ERROR_SLEEP}s then retrying page {page_id}…")
                    time.sleep(SERVER_ERROR_SLEEP)
                    retries += 1
                    continue
                if err_code == 190:
                    sys.exit(
                        "  [!] FATAL: Meta access token is invalid or expired "
                        "(error 190). Renew META_ACCESS_TOKEN in GitHub Secrets."
                    )
                print(f"      [!] API {resp.status_code}: {err_msg}")
                return results, seen_ids, True   # signal: API error
            retries = 0
            data = resp.json()
            for ad in data.get("data", []):
                ad_id  = ad["id"]
                seen_ids.add(ad_id)
                bodies = ad.get("ad_creative_bodies") or []
                spend  = ad.get("spend") or {}
                imp    = ad.get("impressions") or {}
                if bodies:
                    is_removed = any(REMOVAL_TEXT in (b or "").lower() for b in bodies)
                else:
                    # API returned this ad but with no body — Meta strips content on removal
                    is_removed = True
                results[ad_id] = {
                    "is_removed":      is_removed,
                    "spend_min":       spend.get("lower_bound"),
                    "spend_max":       spend.get("upper_bound"),
                    "impressions_min": imp.get("lower_bound"),
                    "impressions_max": imp.get("upper_bound"),
                    "currency":        ad.get("currency"),
                }
            url    = data.get("paging", {}).get("next")
            params = {}
            if url:
                time.sleep(0.3)
    except requests.RequestException as e:
        print(f"      [!] Request failed: {e}")
        return results, seen_ids, True
    return results, seen_ids, False


# ── Core ─────────────────────────────────────────────────────────────────────

def process_db(db_path: str, country: str, args, token: str,
               existing_log_ids: set) -> dict:
    """Check one DB (CY or MT). Returns summary counts."""
    label = country
    print(f"\n{'═'*60}")
    print(f"  {label} — {db_path}")
    print(f"{'═'*60}")

    conn = sqlite3.connect(db_path, timeout=30)
    conn.execute("PRAGMA journal_mode=WAL")

    if args.active_only:
        ads = load_active_page_ads(conn)
        mode_label = "active-only"
    else:
        ads = load_unchecked(conn, only_unchecked=not args.all,
                             recheck_days=args.recheck_days,
                             since=args.since, limit=args.limit)
        mode_label = "full"

    if not ads:
        print(f"  Nothing to check.")
        conn.close()
        return {'total': 0, 'removed': 0, 'active': 0, 'skipped': 0}

    print(f"  Mode         : {mode_label}")
    print(f"  Ads to check : {len(ads):,}")

    # Group by page_id to batch API calls
    by_page: dict[str, list] = defaultdict(list)
    no_page: list = []
    for a in ads:
        pid = a['page_id']
        if pid:
            by_page[pid].append(a)
        else:
            no_page.append(a)

    total_pages = len(by_page)
    print(f"  Pages total  : {total_pages:,}  ({len(no_page)} ads missing page_id — skipped)")

    # ── Max-pages cursor: process a slice of pages per run ───────────────────
    max_pages = getattr(args, 'max_pages', 0)
    if max_pages > 0 and not args.active_only:
        state_file = CHECK_REMOVED_STATE_CY if country == "CY" else CHECK_REMOVED_STATE_MT
        cursor_key = f"check_removed_{country.lower()}_page_cursor"

        # Sort pages deterministically so the cursor is stable across runs
        sorted_page_ids = sorted(by_page.keys())

        # Find the start index from the saved cursor (page_id of first page to process)
        cursor = load_page_cursor(state_file, cursor_key)
        if cursor:
            try:
                start_idx = sorted_page_ids.index(cursor)
            except ValueError:
                # Cursor page no longer in the list — wrap to beginning
                start_idx = 0
        else:
            start_idx = 0

        end_idx = min(start_idx + max_pages, len(sorted_page_ids))
        pages_this_run = sorted_page_ids[start_idx:end_idx]

        # Save cursor for next run: first page of the next chunk (empty = wrap)
        next_cursor = sorted_page_ids[end_idx] if end_idx < len(sorted_page_ids) else ""
        save_page_cursor(state_file, cursor_key, next_cursor)

        remaining = len(sorted_page_ids) - end_idx
        print(f"  This run     : pages {start_idx}–{end_idx - 1} of {len(sorted_page_ids)} "
              f"({remaining} deferred to next run{' — wrapping' if not next_cursor else ''})")

        # Restrict by_page to only the pages for this run
        by_page = {pid: by_page[pid] for pid in pages_this_run}
    # ─────────────────────────────────────────────────────────────────────────

    print(f"  Pages to run : {len(by_page):,}")
    print(f"  Sleep        : {args.sleep}s between pages")
    print("─" * 60)

    now = datetime.now(timezone.utc).isoformat()
    total_removed = 0
    total_active  = 0
    total_skipped = 0
    batch: list[tuple] = []
    newly_removed: list[str] = []   # ad_ids newly detected as removed this run
    BATCH_SIZE = 100
    page_count = 0
    consec_errors = 0              # consecutive API error counter

    # Pre-load which ads in THIS RUN's cursor window are already removed in the DB
    # so we don't log them as "new" when they were previously known.
    # Scoped to by_page (post-cursor-slice) to avoid SQLite's 999-variable limit.
    all_ids = [a['ad_archive_id'] for page_ads in by_page.values() for a in page_ads]
    if all_ids:
        ph = ",".join("?" * len(all_ids))
        already_removed_in_db = set(
            r[0] for r in conn.execute(
                f"SELECT ad_archive_id FROM politician_ads "
                f"WHERE ad_archive_id IN ({ph}) AND removed=1", all_ids
            ).fetchall()
        )
    else:
        already_removed_in_db = set()

    for page_id, page_ads in by_page.items():
        # Use the oldest ad_start_date in this group as the since_date
        dates     = [a['ad_start_date'] for a in page_ads if a['ad_start_date']]
        since_str = min(dates) if dates else "2025-01-01"

        api_results, seen_ids, had_error = fetch_page_ads(page_id, since_str, country, token)
        page_count += 1

        if had_error:
            consec_errors += 1
            if consec_errors >= CONSEC_ERROR_ABORT:
                print(f"\n  [!] {consec_errors} consecutive API errors — Meta API appears to be down.")
                print(f"  [!] Aborting {label} check early to avoid wasting timeout budget.")
                break
            total_skipped += len(page_ads)
            continue
        else:
            consec_errors = 0

        for a in page_ads:
            ad_id = a['ad_archive_id']

            if ad_id in api_results:
                # API returned this ad — use the is_removed flag from body/no-body check
                info       = api_results[ad_id]
                is_removed = info["is_removed"]
                spend_row  = (ad_id, 1 if is_removed else 0, now,
                              info["spend_min"], info["spend_max"],
                              info["impressions_min"], info["impressions_max"],
                              info["currency"])
            elif seen_ids:
                # API call succeeded and returned other ads for this page,
                # but this specific ad ID was absent entirely → removed by Meta
                is_removed = True
                spend_row  = (ad_id, 1, now, None, None, None, None, None)
            else:
                # API returned nothing for this page (empty response) — inconclusive
                total_skipped += 1
                continue

            if is_removed:
                total_removed += 1
                batch.append(spend_row)
                if ad_id not in already_removed_in_db:
                    newly_removed.append(ad_id)
                    print(f"  ⚠ NEW REMOVAL  {ad_id}  (page {page_id})")
                else:
                    print(f"  ⚠ removed      {ad_id}  (already known)")
            else:
                total_active += 1
                batch.append(spend_row)

        if len(batch) >= BATCH_SIZE:
            save_results(conn, batch)
            batch.clear()

        if page_count < len(by_page) and args.sleep > 0:
            time.sleep(args.sleep)

    if batch:
        save_results(conn, batch)

    # Log newly detected removals to Excel
    if newly_removed:
        log_new_removals(conn, newly_removed, country, now, existing_log_ids)

    conn.close()

    total_checked = total_removed + total_active
    print(f"\n  {label} Summary:")
    print(f"    Checked  : {total_checked:,}")
    print(f"    Active   : {total_active:,}")
    print(f"    REMOVED  : {total_removed:,}")
    print(f"    Skipped  : {total_skipped:,}  (ad not returned by API for that page)")

    return {'total': total_checked, 'removed': total_removed,
            'active': total_active, 'skipped': total_skipped}


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="API-based removal detection for CY + MT (no browser required)"
    )
    parser.add_argument('--cy',           action='store_true', help='CY only')
    parser.add_argument('--mt',           action='store_true', help='MT only')
    parser.add_argument('--all',          action='store_true',
                        help='Re-check all active ads (default: unchecked + stale)')
    parser.add_argument('--active-only',  action='store_true',
                        help='Fast mode: only pages with currently-running ads (~78s CY, ~90s MT). '
                             'Safe to run hourly.')
    parser.add_argument('--since',        default='',
                        help='Only check ads with start_date >= DATE')
    parser.add_argument('--limit',        type=int, default=0,
                        help='Max ads per DB (0 = no limit)')
    parser.add_argument('--recheck-days', type=float, default=7,
                        help='Re-check active ads not checked in N days (default: 7, 0=disable)')
    parser.add_argument('--sleep',        type=float, default=1.5,
                        help='Sleep between page API requests in seconds (default: 1.5)')
    parser.add_argument('--max-pages',    type=int, default=0,
                        help='Max pages to check per run; cursor saved to fetch_state so next '
                             'run continues where this one left off (0 = no limit, default: 0). '
                             'Ignored in --active-only mode.')
    args = parser.parse_args()

    cy_token = os.environ.get("META_ACCESS_TOKEN")
    mt_token = os.environ.get("META_ACCESS_TOKEN_MT") or cy_token  # fall back to shared token
    if not cy_token:
        sys.exit("ERROR: META_ACCESS_TOKEN not set in environment or .env")

    run_cy = not args.mt   # run CY unless --mt-only
    run_mt = not args.cy   # run MT unless --cy-only

    # Load existing log IDs once — shared across CY and MT to avoid duplicates
    existing_log_ids = _load_log_ids()

    totals = {'total': 0, 'removed': 0, 'active': 0, 'skipped': 0}

    if run_cy:
        r = process_db(CY_DB, "CY", args, cy_token, existing_log_ids)
        for k in totals:
            totals[k] += r[k]

    if run_mt:
        r = process_db(MT_DB, "MT", args, mt_token, existing_log_ids)
        for k in totals:
            totals[k] += r[k]

    if run_cy and run_mt:
        print(f"\n{'═'*60}")
        print(f"  COMBINED TOTAL")
        print(f"    Checked  : {totals['total']:,}")
        print(f"    Active   : {totals['active']:,}")
        print(f"    REMOVED  : {totals['removed']:,}")
        print(f"    Skipped  : {totals['skipped']:,}")


if __name__ == "__main__":
    main()
