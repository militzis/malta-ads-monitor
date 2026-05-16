"""
make_removed_spend_report.py — Spend analysis on Meta-removed election ads.

Generates removed_spend_YYYY-MM-DD.xlsx with:
  Sheet 1 — Summary         (CY + MT headline figures)
  Sheet 2 — CY by Candidate (removed spend vs total spend, % removed)
  Sheet 3 — CY Removed Ads  (every individual removed ad with spend)
  Sheet 4 — MT by Candidate
  Sheet 5 — MT Removed Ads

Usage:
    python make_removed_spend_report.py
"""

import os, sys, sqlite3
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BASE  = os.path.dirname(os.path.abspath(__file__))
DB_CY = os.path.join(BASE, "politician_ads.db")
DB_MT = os.path.join(BASE, "politician_ads_mt.db")
TODAY = str(date.today())

AD_URL = "https://www.facebook.com/ads/library/?id={}"

# ── Colours ───────────────────────────────────────────────────────────────────
RED_FILL    = PatternFill("solid", fgColor="C00000")
ORANGE_FILL = PatternFill("solid", fgColor="ED7D31")
GREY_FILL   = PatternFill("solid", fgColor="404040")
LTGREY_FILL = PatternFill("solid", fgColor="D9D9D9")
WHITE_FONT  = Font(bold=True, color="FFFFFF")
BOLD        = Font(bold=True)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


# ── Helpers ───────────────────────────────────────────────────────────────────

def write_header(ws, headers, fill=GREY_FILL, font=WHITE_FONT):
    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[ws.max_row].height = 22
    ws.freeze_panes = f"A{ws.max_row + 1}"


def autofit(ws, max_width=52):
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, max_width)


def fmt_name(q):
    return (q or "").split("|")[0].strip()

def fmt_party(q):
    parts = (q or "").split("|")
    return parts[1].strip() if len(parts) > 1 else ""

def fmt_district(q):
    parts = (q or "").split("|")
    return parts[2].strip() if len(parts) > 2 else ""

def midpoint(lo, hi):
    lo = lo or 0
    hi = hi or 0
    return (lo + hi) / 2


# ── Queries ───────────────────────────────────────────────────────────────────

def by_candidate(conn):
    return conn.execute("""
        WITH rem AS (
            SELECT politician_query, party,
                   COUNT(*)           AS rem_ads,
                   SUM(spend_min)     AS rem_smin,
                   SUM(spend_max)     AS rem_smax,
                   SUM((COALESCE(spend_min,0)+COALESCE(spend_max,0))/2.0) AS rem_mid
            FROM politician_ads
            WHERE removed=1 AND election_related='YES'
            GROUP BY politician_query
        ),
        tot AS (
            SELECT politician_query,
                   COUNT(*)           AS tot_ads,
                   SUM(spend_max)     AS tot_smax,
                   SUM((COALESCE(spend_min,0)+COALESCE(spend_max,0))/2.0) AS tot_mid
            FROM politician_ads
            WHERE election_related='YES'
            GROUP BY politician_query
        )
        SELECT r.politician_query, r.party,
               r.rem_ads, r.rem_smin, r.rem_smax, r.rem_mid,
               t.tot_ads, t.tot_smax, t.tot_mid,
               ROUND(100.0 * r.rem_smax / NULLIF(t.tot_smax, 0), 1) AS pct_spend_removed
        FROM rem r JOIN tot t USING (politician_query)
        ORDER BY r.rem_smax DESC
    """).fetchall()


def all_removed_ads(conn):
    return conn.execute("""
        SELECT ad_archive_id, politician_query, party,
               page_name, ad_start_date, ad_stop_date,
               spend_min, spend_max, impressions_min, impressions_max,
               currency, removed_checked_at
        FROM politician_ads
        WHERE removed=1 AND election_related='YES'
        ORDER BY spend_max DESC NULLS LAST, ad_start_date DESC
    """).fetchall()


def db_summary(conn):
    # Use CASE WHEN instead of FILTER — FILTER was added in SQLite 3.30 (2019)
    # and is not available on all platforms (e.g. older Windows SQLite DLLs).
    return conn.execute("""
        SELECT
            SUM(CASE WHEN removed=1 AND election_related='YES' THEN 1 ELSE 0 END)          AS removed_ads,
            SUM(CASE WHEN election_related='YES'               THEN 1 ELSE 0 END)          AS total_ads,
            COUNT(DISTINCT CASE WHEN removed=1 AND election_related='YES'
                                THEN politician_query END)                                  AS removed_candidates,
            SUM(CASE WHEN removed=1 AND election_related='YES' THEN spend_min  ELSE NULL END) AS rem_smin,
            SUM(CASE WHEN removed=1 AND election_related='YES' THEN spend_max  ELSE NULL END) AS rem_smax,
            SUM(CASE WHEN election_related='YES'               THEN spend_max  ELSE NULL END) AS tot_smax
        FROM politician_ads
    """).fetchone()


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_summary_sheet(ws, cy_sum, mt_sum):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    def section(label, s, fill):
        ws.append([])
        r = ws.max_row + 1
        ws.append([label])
        title_cell = ws.cell(r, 1)
        title_cell.font = Font(bold=True, size=13, color="FFFFFF")
        title_cell.fill = fill
        title_cell.alignment = Alignment(horizontal="left")
        ws.row_dimensions[r].height = 20
        ws.merge_cells(f"A{r}:D{r}")

        rows = [
            ("Removed ads",              s[0],
             f"{s[0]:,} of {s[1]:,} total ({s[0]/s[1]*100:.0f}%)" if s[1] else "—"),
            ("Candidates with removals", s[2], f"{s[2]:,}"),
            ("Spend on removed (min €)", s[3], f"€{s[3]:,.0f}" if s[3] else "€0"),
            ("Spend on removed (max €)", s[4], f"€{s[4]:,.0f}" if s[4] else "€0"),
            ("% of total spend removed", None,
             f"{s[4]/s[5]*100:.1f}%" if (s[5] and s[4] is not None) else "—"),
        ]
        for label2, _, display in rows:
            ws.append(["", label2, display])
            row = ws[ws.max_row]
            row[1].font = BOLD
            row[2].alignment = Alignment(horizontal="right")

    build_summary_sheet.cy_fill = RED_FILL
    build_summary_sheet.mt_fill = ORANGE_FILL

    ws.append(["Removed Ads — Spend Analysis", "", f"Generated {TODAY}"])
    ws[ws.max_row][0].font = Font(bold=True, size=15)
    ws[ws.max_row][2].font = Font(italic=True, color="808080")

    section("CYPRUS (CY)", cy_sum, RED_FILL)
    section("MALTA (MT)",  mt_sum, ORANGE_FILL)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 28


def build_candidate_sheet(ws, title, rows, fill):
    ws.title = title
    ws.sheet_view.showGridLines = False

    headers = [
        "Candidate", "Party", "District",
        "Removed Ads", "Removed Spend Min (€)", "Removed Spend Max (€)",
        "Total Ads", "Total Spend Max (€)", "% Spend Removed",
    ]
    write_header(ws, headers, fill=fill)

    for r in rows:
        cand     = fmt_name(r[0])
        party    = r[1] or fmt_party(r[0])
        district = fmt_district(r[0])
        rem_ads, rem_smin, rem_smax = r[2], r[3] or 0, r[4] or 0
        tot_ads, tot_smax           = r[6], r[7] or 0
        pct = r[9]

        row_data = [cand, party, district,
                    rem_ads, rem_smin, rem_smax,
                    tot_ads, tot_smax,
                    (pct / 100) if pct else None]
        ws.append(row_data)

        data_row = ws[ws.max_row]
        # Highlight candidates where >80% of spend was removed
        if pct and pct >= 80:
            for cell in data_row:
                cell.fill = PatternFill("solid", fgColor="FFE0E0")
        # % column as percentage format
        pct_cell = data_row[8]
        if pct_cell.value is not None:
            pct_cell.number_format = "0.0%"
        # € columns
        for idx in (4, 5, 7):
            data_row[idx].number_format = '#,##0'

    autofit(ws)
    # Fixed widths for key columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["I"].width = 14


def build_ads_sheet(ws, title, rows, fill):
    ws.title = title
    ws.sheet_view.showGridLines = False

    headers = [
        "Ad ID", "Candidate", "Party", "Page Name",
        "Start Date", "End Date",
        "Spend Min (€)", "Spend Max (€)",
        "Impressions Min", "Impressions Max",
        "Currency", "Checked At", "Ad Library URL",
    ]
    write_header(ws, headers, fill=fill)

    for r in rows:
        ad_id = r[0]
        ws.append([
            ad_id,
            fmt_name(r[1]),
            r[2] or fmt_party(r[1]),
            r[3],
            r[4], r[5],
            r[6] or 0, r[7] or 0,
            r[8], r[9],
            r[10],
            (r[11] or "")[:19],
            AD_URL.format(ad_id),
        ])
        data_row = ws[ws.max_row]
        for idx in (6, 7):
            data_row[idx].number_format = '#,##0'
        # URL as hyperlink-style text
        url_cell = data_row[12]
        url_cell.font = Font(color="0563C1", underline="single")

    autofit(ws)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["M"].width = 55


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    cy = sqlite3.connect(DB_CY)
    mt = sqlite3.connect(DB_MT)

    cy_cands  = by_candidate(cy)
    mt_cands  = by_candidate(mt)
    cy_ads    = all_removed_ads(cy)
    mt_ads    = all_removed_ads(mt)
    cy_sum    = db_summary(cy)
    mt_sum    = db_summary(mt)

    cy.close()
    mt.close()

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    ws_sum = wb.create_sheet("Summary")
    build_summary_sheet(ws_sum, cy_sum, mt_sum)

    ws_cy = wb.create_sheet("CY — By Candidate")
    build_candidate_sheet(ws_cy, "CY — By Candidate", cy_cands, RED_FILL)

    ws_cy_ads = wb.create_sheet("CY — Removed Ads")
    build_ads_sheet(ws_cy_ads, "CY — Removed Ads", cy_ads, RED_FILL)

    ws_mt = wb.create_sheet("MT — By Candidate")
    build_candidate_sheet(ws_mt, "MT — By Candidate", mt_cands, ORANGE_FILL)

    ws_mt_ads = wb.create_sheet("MT — Removed Ads")
    build_ads_sheet(ws_mt_ads, "MT — Removed Ads", mt_ads, ORANGE_FILL)

    out = os.path.join(BASE, f"removed_spend_{TODAY}.xlsx")
    wb.save(out)
    print(f"Saved: {out}")
    print(f"  CY: {len(cy_cands)} candidates, {len(cy_ads)} removed ads")
    print(f"  MT: {len(mt_cands)} candidates, {len(mt_ads)} removed ads")


if __name__ == "__main__":
    main()
