"""
Build candidate_ads_combined_mt.xlsx from politician_ads_mt.db.
Malta-specific — no Greek/Latin split, uses page_blocklist_mt.json.
"""

import sys
import sqlite3
import json
import os
from collections import defaultdict

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import get_column_letter

BASE      = os.path.dirname(os.path.abspath(__file__))
DB_PATH   = os.path.join(BASE, "politician_ads_mt.db")
BL_FILE   = os.path.join(BASE, "page_blocklist_mt.json")
OUT_FILE  = os.path.join(BASE, "candidate_ads_combined_mt.xlsx")


# ── Blocklist ──────────────────────────────────────────────────────────────────
def load_blocklist():
    if os.path.exists(BL_FILE):
        with open(BL_FILE, encoding='utf-8') as f:
            return set(json.load(f).get('pages', {}).keys())
    return set()

BLOCKLIST = load_blocklist()
print(f"Loaded {len(BLOCKLIST)} blocked page IDs")


# ── Columns ────────────────────────────────────────────────────────────────────
AD_COLUMNS = [
    "Candidate", "Party", "District",
    "Page Name", "Page ID", "Bylines",
    "Start Date", "Stop Date",
    "Impressions Min", "Impressions Max",
    "Spend Min", "Spend Max", "Currency",
    "View Ad", "Checked At", "Removed",
]

SUM_COLUMNS = [
    "Candidate", "Party", "District",
    "Total Ads", "Active Ads", "Inactive Ads", "Removed Ads",
    "Impressions Min (total)", "Impressions Max (total)",
    "Spend Min (total)", "Spend Max (total)", "Currency",
    "Unique Pages",
    "Ad Library Link",
]


# ── Load DB ────────────────────────────────────────────────────────────────────
print("Loading DB …")
conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
all_rows_raw = conn.execute(
    "SELECT * FROM politician_ads WHERE ad_start_date >= '2025-10-01'"
).fetchall()
conn.close()

all_rows = [dict(r) for r in all_rows_raw]
print(f"  {len(all_rows):,} rows loaded")

# Apply blocklist (double-check — should already be clean)
filtered = [r for r in all_rows if str(r.get("page_id") or "") not in BLOCKLIST]
skipped = len(all_rows) - len(filtered)
if skipped:
    print(f"  Blocklist filtered {skipped} rows")
print(f"  {len(filtered):,} rows after filter")


# ── Build Excel ────────────────────────────────────────────────────────────────
print("Writing Excel …")
wb = Workbook()
ws = wb.active
ws.title = "Combined Ads"

header_font = Font(bold=True)

# Header
for col_idx, col_name in enumerate(AD_COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

# Data rows
for r in filtered:
    query     = r.get("politician_query", "")
    candidate = query.split("|")[0]
    party     = query.split("|")[1] if "|" in query else r.get("party", "")
    district  = query.split("|")[2] if query.count("|") >= 2 else r.get("district", "")
    ad_id     = r.get("ad_archive_id") or ""
    public_url = f"https://www.facebook.com/ads/library/?id={ad_id}" if ad_id else ""

    removed = r.get("removed")
    removed_label = "YES" if removed == 1 else ""

    row_data = [
        candidate,
        party or r.get("party"),
        district or r.get("district"),
        r.get("page_name"),
        r.get("page_id"),
        r.get("bylines"),
        r.get("ad_start_date"),
        r.get("ad_stop_date"),
        r.get("impressions_min"),
        r.get("impressions_max"),
        r.get("spend_min"),
        r.get("spend_max"),
        r.get("currency"),
        "View Ad",
        r.get("checked_at"),
        removed_label,
    ]

    row_idx = ws.max_row + 1
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        if col_idx == 14 and public_url:   # "View Ad" column
            cell.hyperlink = public_url
            cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="center")

# Column widths
col_widths = [30, 10, 12, 40, 20, 20, 12, 12, 14, 14, 10, 10, 8, 10, 28, 9]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"


# ── Summary sheet ──────────────────────────────────────────────────────────────
print("Writing Summary sheet …")
ws2 = wb.create_sheet(title="Summary")

for col_idx, col_name in enumerate(SUM_COLUMNS, 1):
    cell = ws2.cell(row=1, column=col_idx, value=col_name)
    cell.font = header_font

# Aggregate per candidate
cand_data: dict = defaultdict(lambda: {
    "party": "", "district": "",
    "total": 0, "active": 0, "inactive": 0, "removed": 0,
    "impr_min": 0, "impr_max": 0,
    "spend_min": 0, "spend_max": 0,
    "currency": "",
    "pages": set(),
})

for r in filtered:
    query = r.get("politician_query", "")
    name  = query.split("|")[0]
    party = query.split("|")[1] if "|" in query else r.get("party", "")
    district = query.split("|")[2] if query.count("|") >= 2 else r.get("district", "")

    d = cand_data[name]
    d["party"]    = party or d["party"]
    d["district"] = district or d["district"]
    d["total"]   += 1
    is_removed = r.get("removed") == 1
    has_stopped = bool(r.get("ad_stop_date"))
    if is_removed:
        d["removed"] += 1
    elif has_stopped:
        d["inactive"] += 1
    else:
        d["active"]  += 1
    d["impr_min"]  += r.get("impressions_min") or 0
    d["impr_max"]  += r.get("impressions_max") or 0
    d["spend_min"] += r.get("spend_min") or 0
    d["spend_max"] += r.get("spend_max") or 0
    if r.get("currency"):
        d["currency"] = r["currency"]
    pid = str(r.get("page_id") or "").strip()
    if pid and pid != "0":
        d["pages"].add(pid)

# Sort by total ads descending
sorted_cands = sorted(cand_data.items(), key=lambda x: x[1]["total"], reverse=True)

right = Alignment(horizontal="right")

for name, d in sorted_cands:
    # Build Ad Library link using the first page ID found
    page_id = next(iter(d["pages"]), None)
    ad_lib_url = (
        f"https://www.facebook.com/ads/library/?active_status=all&ad_type=all"
        f"&country=MT&media_type=all&view_all_page_id={page_id}"
    ) if page_id else ""

    row_idx = ws2.max_row + 1
    row_data = [
        name,
        d["party"],
        d["district"] or None,
        d["total"],
        d["active"],
        d["inactive"],
        d["removed"] or None,
        d["impr_min"] or None,
        d["impr_max"] or None,
        d["spend_min"] or None,
        d["spend_max"] or None,
        d["currency"] or None,
        len(d["pages"]),
        "Ad Library" if ad_lib_url else None,
    ]

    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if col_idx >= 4:
            cell.alignment = right
        # Ad Library hyperlink
        if col_idx == 13 and ad_lib_url:
            cell.hyperlink = ad_lib_url
            cell.font = Font(color="0563C1", underline="single")
            cell.alignment = Alignment(horizontal="center")

ws2.freeze_panes = "A2"
sum_widths = [30, 10, 12, 11, 11, 11, 11, 22, 22, 18, 18, 9, 13, 14]
for i, w in enumerate(sum_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT_FILE)
print(f"\nDone! Saved to: {OUT_FILE}")
print(f"Total ad rows: {len(filtered):,}  |  Candidates in summary: {len(cand_data):,}")
print(f"Candidates with ads: {sum(1 for d in cand_data.values() if d['total'] > 0)}")
