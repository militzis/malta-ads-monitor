"""For each advertiser, dump the 3 most-recent ads with:
   - clickable TikTok library URL (always works)
   - local file path (for instant playback if downloaded)
   - transcript snippet
   - reviewer-decision column for you to fill in
"""
import sys, sqlite3, json, os, glob
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

DB        = r'C:\Users\milit\meta_pipeline_data\politician_ads.db'
CREATIVES = r'C:\Users\milit\meta_pipeline_data\creatives'
OUT       = r"C:\Users\milit\Downloads\tiktok_verify_recent_ads.xlsx"

c = sqlite3.connect(DB)

# Same category labels as the profiles Excel
ANNOTATIONS = {
    # Original 7 confirmed
    '7578569963879792657': ('Candidate', 'Παζάρος Χαράλαμπος', 'ΔΗΣΥ'),
    '7488289249494679569': ('Candidate', 'Ιωάννου Κλεονίκη', 'ΑΜΔΗ'),
    '7563644604046852097': ('Candidate', 'Χρυσάνθου Λοΐζος', 'ΑΜΔΗ'),
    '7533251756613189648': ('Candidate', 'Φλουρέντζου Μάριος', 'ΕΛΑΜ'),
    '7481250791723008017': ('Candidate', 'Ηλιά Μάριος', 'ΔΗΣΥ'),
    '7554065792900448257': ('Candidate', 'Καλοπαίδης Μιχάλης', 'ΒΟΛΤ'),
    '7612669198808039440': ('Candidate', 'Πουλλικκάς Μάριος', 'ΕΛΑΜ'),
    # Concat-match confirmed
    '7603797428927578113': ('Candidate', 'Παστέλλα Μαρία', 'ΕΛΑΜ'),
    '7604116497765711889': ('Candidate', 'Πετρίδου Στέλλα', 'ΟΙΚΟΛΟΓΟΙ'),
    '7579460984457150480': ('Candidate', 'Αθανασίου Μαριάννα', 'ΛΑΚΕΔΑΙΜΟΝΙΟΙ'),
    # Party / podcast / news / commentator / unverified
    '7628713809988960272': ('Party account',    '@almalimassol — ΑΛΜΑ regional', 'ΑΛΜΑ'),
    '7450088911596011536': ('Party coordinator','@nakiskyriakou — ΣΗΚΟΥ ΠΑΝΩ slogan', 'ΣΗΚΟΥ ΠΑΝΩ'),
    '7600116160427851793': ('News outlet',      '@tanea.cy', '-'),
    '7464520683175804929': ('Podcast',          '@tolkerscy', '-'),
    '7444996016945774609': ('Commentator',      '@gnosths_ths_istorias', '-'),
    '7479401042212487184': ('Satirist',         '@gastrimargos', '-'),
    '7454901333884141584': ('Unverified',       '@angelosiacovides', '?'),
}
# Add concat-match candidates by handle
HANDLE_TO_CAND = {
    'argentoulaioannou':   ('Candidate', 'Ιωάννου Αργεντούλα', 'ΑΚΕΛ'),
    'argyrosevangelou':    ('Candidate', 'Ευαγγέλου Αργυρός', 'ΔΗΣΥ'),
    'chrysanthossavvidis': ('Candidate', 'Σαββίδης Χρύσανθος', 'ΔΗΚΟ'),
    'kyprianouanna':       ('Candidate', 'Κυπριανού Άννα', 'ΔΗΠΑ'),
    'kyproskyprianou4':    ('Candidate', 'Κυπριανού Κύπρος', 'ΑΛΜΑ'),
    'lakiskonstantinou1':  ('Candidate', 'Κωνσταντίνου Λάκης', 'ΕΔΕΚ'),
    'pamposkiskonstantinos':('Candidate','Παμπόσκης Κωνσταντίνος', 'ΕΛΑΜ'),
    'petros.minas2':       ('Candidate', 'Μηνάς Πέτρος', 'ΔΕΚ'),
    'petrouiakovos':       ('Candidate', 'Πέτρου Ιάκωβος', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ'),
    'ploutarchosparris':   ('Candidate', 'Παρρής Πλούταρχος', 'ΣΗΚΟΥ ΠΑΝΩ'),
    'theodosisavgousti':   ('Candidate', 'Αυγουστή Θεοδόσης', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ'),
    'antonis.antoniou_':   ('Candidate', 'Αντωνή Αντώνης', 'ΕΔΕΚ'),
    # Post-resolution candidates (added today)
    'marios_stavrou_':     ('Candidate', 'Σταύρου Μάριος',   'ΑΜΔΗ'),
    'parismarkou':         ('Candidate', 'Μάρκου Πάρις',     'ΔΗΣΥ'),
    'michalis_fellas':     ('Candidate', 'Φελλάς Μιχάλης',   'ΔΗΣΥ'),
    'marios.neofytou':     ('Candidate', 'Νεοφύτου Μάριος',  'ΑΛΜΑ'),
    'elenachristou1':      ('Candidate', 'Χρίστου Έλεν',     'ΣΗΚΟΥ ΠΑΝΩ'),
    # User-verified manually
    'deme2023':            ('Candidate', 'Χατζησταύρου Δήμητρα', 'ΑΜΔΗ'),
    'theanicolaou6':       ('Candidate', 'Νικολάου Θέα',          'ΔΗΠΑ'),
    'evgenioshamboullas':  ('Candidate', 'Χαμπούλλας Ευγένιος',   'ΕΛΑΜ'),
    'paliosn':             ('Candidate', 'Πάλιος Νίκος',           'ΑΜΔΗ'),
    'elena.vrahimi':       ('Candidate', 'Βραχίμη Έλενα',          'ΔΗΠΑ'),
    'apostolouaa':         ('Needs verification', 'Αποστόλου (homonym)', 'ΛΑΕ_or_ΔΗΚΟ'),
    'phivos.doukanaris':   ('Candidate', 'Δουκανάρης Φοίβος',      'ΑΜΔΗ'),
    'steliosmohicanstylianou': ('Candidate', 'Στυλιανού Στέλιος (Κασιουλή)', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ'),
    'apostolouaa':           ('Candidate', 'Αποστόλου Ανδρέας',     'ΔΗΚΟ'),
    'maria.loizou85':        ('Candidate', 'Λοΐζου Μαρία',           'ΑΜΔΗ'),
    'adrianachristodoulou13':('Candidate', 'Χριστοδούλου Αδριάνα',   'ΕΛΑΜ'),
    'lysandrides':           ('Candidate', 'Λυσανδρίδης Γεώργιος',   'ΔΗΣΥ'),
    'steliosstylianou78':    ('Candidate', 'Στυλιανού Στέλιος',       'ΕΛΑΜ'),
    'demetris.aletraris':    ('Candidate', 'Αλετράρης Δημήτρης',     'ΑΜΔΗ'),
    # Bio-scan / user-verified (2026-05-18)
    'allasoneagian':         ('Candidate', 'Παπαχριστοφόρου Στυλιανός', 'ΑΜΔΗ'),
    'andreas_papacharalambous':('Candidate','Παπαχαραλάμπους Ανδρέας', 'ΕΛΑΜ'),
    'florentzos_karayiannas':('Candidate', 'Καραγιάννας Φλωρέντζος',   'ΔΗΚΟ'),
    'demosg_cy':             ('Candidate', 'Γεωργιάδης Δήμος',          'ΔΗΣΥ'),
    'stamos.papavasili':     ('Candidate', 'Παπαβασιλείου Σταμάτης',    'ΟΙΚΟΛΟΓΟΙ'),
    'modecristo':            ('Candidate', 'Μοδέστου Χρίστος',          'ΟΙΚΟΛΟΓΟΙ'),
    'tasos2026':             ('Candidate', 'Αναστασίου Αναστάσιος',     'ΑΜΔΗ'),
    'ari.habeshian':         ('Candidate', 'Χαπεσιάν Άρι',              'ΔΗΠΑ'),
    'paraschou1':            ('Candidate', 'Παρασχού Αντώνης',          'ΣΗΚΟΥ ΠΑΝΩ'),
    'theofilos891':          ('Party supporter', 'Θεόφιλος — ΕΛΑΜ supporter', 'ΕΛΑΜ'),
    'tsitsis_channel':       ('Candidate', 'Τσίτσης Γιάννης',           'ΑΜΔΗ'),
    'fotinitsiridou':        ('Candidate', 'Τσιρίδου Φωτεινή',          'ΔΗΣΥ'),
    'charis_av':             ('Party supporter', 'Χάρης — ΣΗΚΟΥ ΠΑΝΩ supporter', 'ΣΗΚΟΥ ΠΑΝΩ'),
    'blackdreamscottage':    ('Party supporter', 'ΑΜΔΗ supporter', 'ΑΜΔΗ'),
    'phedonphedonos':        ('Politician', 'Φαίδων Φαίδωνος — Δήμαρχος Πάφου', 'ΔΗΣΥ'),
    'adiafthoroi':           ('Commentator', 'Αδιάφθοροι — anti-small-party account', '-'),
    'neadynami':             ('Political movement', 'Νέα Δυναμική', 'Νέα Δυναμική'),
    'elena_stefani19':       ('Party supporter', 'Έλενα Στεφανή — ΕΛΑΜ supporter', 'ΕΛΑΜ'),
    'yiannis.laouris':       ('Needs verification', 'Γιάννης Λαουρής — party TBD', '?'),
    'andreamanoli7':         ('Candidate', 'Θεολόγου Μανωλή Άνδρεα', 'ΔΗΣΥ'),
    'kyriakimanousaki':      ('Candidate', 'Μανουσάκη Κυριακή', 'ΕΔΕΚ'),
    'veronikapavlidou':      ('Candidate', 'Παυλίδου Βερόνικα', 'ΔΕΚ'),
    'voulakokkinou7':        ('Candidate', 'Κοκκίνου Παρασκευή ("Βούλα")', 'ΑΛΜΑ'),
    # Homonyms that need verification
    'steliosmohicanstylianou': ('Candidate', 'Στυλιανού Στέλιος ("O Souvlakis")', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ'),
    'steliosstylianou78':      ('Candidate', 'Στυλιανού Στέλιος',           'ΕΛΑΜ'),
    'maria.loizou85':          ('Candidate', 'Λοΐζου Μαρία',                'ΑΜΔΗ'),
}

def categorize(bid, handle):
    if str(bid) in ANNOTATIONS:
        return ANNOTATIONS[str(bid)]
    if (handle or '').lower() in HANDLE_TO_CAND:
        return HANDLE_TO_CAND[handle.lower()]
    return ('?', '?', '?')

# Get the universe: every advertiser with at least 1 ad
advertisers = list(c.execute("""
    SELECT advertiser_id, advertiser_disclosed_name,
           COUNT(*) AS ads,
           MIN(first_shown), MAX(last_shown)
    FROM tiktok_ads
    GROUP BY advertiser_id, advertiser_disclosed_name
"""))

# Keep only those we care about for verification: known candidates + party/coord/news/etc + unverified
# (Skip the 335 content_keyword noise — those need different triage)
VERIFY_SET = set(ANNOTATIONS.keys())
ADVS = [(bid, h, ads, first, last) for (bid, h, ads, first, last) in advertisers
        if str(bid) in VERIFY_SET or (h or '').lower() in HANDLE_TO_CAND]

print(f"Building verification list for {len(ADVS)} advertisers")

wb = Workbook()
ws = wb.active
ws.title = 'Verify recent ads'

HEADERS = ['Category','Handle','Profile URL','Identified as','Party','# total ads',
           '#','Ad date','Library URL','Local file (play)','Transcript snippet','REVIEWER DECISION']
HEADER_FILL = PatternFill('solid', fgColor='305496')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
HYPER_FONT  = Font(color='0563C1', underline='single')
FILLS = {
    'Candidate': PatternFill('solid', fgColor='C6E0B4'),
    'Party account': PatternFill('solid', fgColor='BDD7EE'),
    'Party coordinator': PatternFill('solid', fgColor='BDD7EE'),
    'News outlet': PatternFill('solid', fgColor='FFF2CC'),
    'Podcast': PatternFill('solid', fgColor='FFF2CC'),
    'Commentator': PatternFill('solid', fgColor='FFF2CC'),
    'Satirist': PatternFill('solid', fgColor='FFF2CC'),
    'Unverified': PatternFill('solid', fgColor='FCE4D6'),
    '?': PatternFill('solid', fgColor='FFFFFF'),
}

for ci, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT

# Sort: candidates first, then party/etc, then unverified
SORT_ORDER = ['Candidate','Party account','Party coordinator','News outlet','Podcast',
              'Commentator','Satirist','Unverified','?']

def sk(adv):
    bid, h, ads, _, _ = adv
    cat, _, _ = categorize(bid, h)
    try: idx = SORT_ORDER.index(cat)
    except ValueError: idx = 999
    return (idx, -ads, str(h))
ADVS.sort(key=sk)

r = 2
for bid, handle, ads, first_all, last_all in ADVS:
    cat, name, party = categorize(bid, handle)
    profile_url = f"https://www.tiktok.com/@{handle}" if handle and not str(handle).isdigit() else ""
    # Pull the 3 most-recent ads for this advertiser (most recent = highest last_shown)
    recent = list(c.execute("""
        SELECT ad_id, first_shown, last_shown, videos_json, image_urls_json,
               transcript, ad_url
        FROM tiktok_ads WHERE advertiser_id=?
        ORDER BY last_shown DESC, first_shown DESC
        LIMIT 3
    """, (bid,)))
    fill = FILLS.get(cat, FILLS['?'])
    for i, (ad_id, fs, ls, vids_j, imgs_j, transcript, ad_url) in enumerate(recent, 1):
        # Build local file path
        sub = os.path.join(CREATIVES, handle or f"bid_{bid}")
        local_file = ''
        if os.path.isdir(sub):
            matches = glob.glob(os.path.join(sub, f"{ad_id}*"))
            if matches:
                local_file = matches[0]
        library_url = f"https://library.tiktok.com/ads/detail/?ad_id={ad_id}"
        snippet = (transcript or '')[:300]
        if transcript and len(transcript) > 300:
            snippet += '...'

        vals = [cat, handle if i==1 else '', profile_url if i==1 else '',
                name if i==1 else '', party if i==1 else '', ads if i==1 else '',
                i, ls or '', library_url, local_file, snippet, '']
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=v)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        # Hyperlinks
        if i==1 and profile_url:
            pc = ws.cell(row=r, column=3); pc.hyperlink = profile_url; pc.font = HYPER_FONT
        lc = ws.cell(row=r, column=9); lc.hyperlink = library_url; lc.value = 'library'; lc.font = HYPER_FONT
        if local_file:
            fc = ws.cell(row=r, column=10); fc.hyperlink = local_file; fc.value = 'play locally'; fc.font = HYPER_FONT
        r += 1

# Column widths
WIDTHS = {'Category':18,'Handle':24,'Profile URL':30,'Identified as':28,'Party':14,
          '# total ads':8,'#':3,'Ad date':12,'Library URL':12,'Local file (play)':14,
          'Transcript snippet':80,'REVIEWER DECISION':22}
for i, h in enumerate(HEADERS, 1):
    ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(h, 14)
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
for row in range(2, ws.max_row + 1):
    ws.row_dimensions[row].height = 55

try:
    wb.save(OUT)
    print(f"\nSaved {OUT}")
except PermissionError:
    from datetime import datetime
    alt = OUT.replace('.xlsx', f"_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(alt)
    print(f"[locked] Saved {alt}")
