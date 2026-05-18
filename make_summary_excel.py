"""
make_summary_excel.py — Daily summary Excel for Cyprus & Malta ad monitoring.

Generates summary_YYYY-MM-DD.xlsx with:
  Sheet 1 — CY Party Overview      (active / inactive / removed / spend / impressions)
  Sheet 2 — CY Top by Spend        (top 30 politicians by max spend)
  Sheet 3 — CY Top by Impressions  (top 30 by max impressions)
  Sheet 4 — CY Active Ads          (all currently running election ads)
  Sheet 5 — CY Removed Ads         (all confirmed removed election ads)
  Sheet 6 — MT Party Overview
  Sheet 7 — MT Active Ads

Usage:
    python make_summary_excel.py
    python make_summary_excel.py --out /path/to/folder
"""

import os, sys, sqlite3, argparse
from datetime import date, datetime

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

BASE       = os.path.dirname(os.path.abspath(__file__))
DB_CY      = os.path.join(BASE, "politician_ads.db")
DB_MT      = os.path.join(BASE, "politician_ads_mt.db")
AD_URL     = "https://www.facebook.com/ads/library/?id={}"


# ── db migration ─────────────────────────────────────────────────────────────

def migrate_db(conn):
    """Add election_related column if missing (needed for Malta DB)."""
    cols = {r[1] for r in conn.execute("PRAGMA table_info(politician_ads)").fetchall()}
    if "election_related" not in cols:
        conn.execute("ALTER TABLE politician_ads ADD COLUMN election_related TEXT")
        conn.commit()


# ── helpers ───────────────────────────────────────────────────────────────────

def fmt_name(q):
    parts = (q or "").split("|")
    return parts[0].strip()

def fmt_range(lo, hi, currency=""):
    if lo is None and hi is None:
        return ""
    lo = lo or 0
    hi = hi or 0
    cur = f" {currency}" if currency else ""
    return f"{lo:,}–{hi:,}{cur}"

def col_letter(n):
    """1-based column index to Excel letter."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


# ── styling ───────────────────────────────────────────────────────────────────

def write_header(ws, headers, fill, font):
    from openpyxl.styles import Alignment
    ws.append(headers)
    hdr_row = ws.max_row   # always style the row just appended, not ws[1]
    for cell in ws[hdr_row]:
        cell.fill  = fill
        cell.font  = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 22
    ws.freeze_panes = f"A{hdr_row + 1}"


def autofit(ws, max_width=55):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 2, max_width)


# ── data queries ──────────────────────────────────────────────────────────────

def _exec(conn, sql, params=()):
    """Execute sql, auto-migrating if election_related column is missing."""
    try:
        return conn.execute(sql, params).fetchall()
    except sqlite3.OperationalError as e:
        if "election_related" in str(e):
            migrate_db(conn)
            return conn.execute(sql, params).fetchall()
        raise


def party_overview(conn, today):
    sql = """
        SELECT
            COALESCE(party,'Unknown')                                         AS party,
            SUM(CASE WHEN removed=0
                 AND (ad_stop_date IS NULL OR ad_stop_date='' OR ad_stop_date>=?)
                 THEN 1 ELSE 0 END)                                           AS active,
            SUM(CASE WHEN removed=0
                 AND ad_stop_date < ? AND ad_stop_date!=''
                 THEN 1 ELSE 0 END)                                           AS inactive,
            SUM(CASE WHEN removed=1 THEN 1 ELSE 0 END)                        AS removed,
            COUNT(*)                                                           AS total,
            SUM(CASE WHEN removed=0 THEN COALESCE(impressions_max,0) ELSE 0 END) AS imp,
            SUM(CASE WHEN removed=0 THEN COALESCE(spend_max,0) ELSE 0 END)    AS spend
        FROM politician_ads
        WHERE election_related='YES'
          AND ad_start_date >= '2025-10-01'
        GROUP BY party
        ORDER BY total DESC
    """
    return _exec(conn, sql, (today, today))


def top_politicians(conn, today, order_col, limit=30):
    sql = f"""
        SELECT
            politician_query,
            party,
            page_name,
            SUM(CASE WHEN removed=0
                 AND (ad_stop_date IS NULL OR ad_stop_date='' OR ad_stop_date>=?)
                 THEN 1 ELSE 0 END)                                           AS active,
            SUM(CASE WHEN removed=0 AND ad_stop_date < ? AND ad_stop_date!=''
                 THEN 1 ELSE 0 END)                                           AS inactive,
            SUM(CASE WHEN removed=1 THEN 1 ELSE 0 END)                        AS removed_cnt,
            MAX(COALESCE(spend_max, 0))                                        AS max_spend,
            MAX(COALESCE(impressions_max, 0))                                  AS max_imp,
            MAX(currency)                                                      AS currency,
            SUM(CASE WHEN removed=0 THEN COALESCE(spend_max,0) ELSE 0 END)    AS total_spend,
            SUM(CASE WHEN removed=0 THEN COALESCE(impressions_max,0) ELSE 0 END) AS total_imp
        FROM politician_ads
        WHERE election_related='YES'
          AND ad_start_date >= '2025-10-01'
        GROUP BY politician_query
        ORDER BY {order_col} DESC
        LIMIT {limit}
    """
    return _exec(conn, sql, (today, today))


def active_ads(conn, today):
    sql = """
        SELECT
            ad_archive_id, politician_query, party, district,
            page_name, ad_start_date,
            impressions_min, impressions_max,
            spend_min, spend_max, currency
        FROM politician_ads
        WHERE election_related='YES'
          AND removed=0
          AND (ad_stop_date IS NULL OR ad_stop_date='' OR ad_stop_date>=?)
          AND ad_start_date >= '2025-10-01'
        ORDER BY ad_start_date DESC
    """
    return _exec(conn, sql, (today,))


def removed_ads(conn):
    sql = """
        SELECT
            ad_archive_id, politician_query, party, district,
            page_name, ad_start_date, removed_checked_at,
            impressions_min, impressions_max,
            spend_min, spend_max, currency
        FROM politician_ads
        WHERE election_related='YES'
          AND removed=1
          AND ad_start_date >= '2025-10-01'
        ORDER BY removed_checked_at DESC
    """
    return _exec(conn, sql)


# ── sheet builders ────────────────────────────────────────────────────────────

def build_party_sheet(ws, rows, country, fill, hdr_font, alt_fill, link_font, total_font):
    from openpyxl.styles import Font, Alignment, PatternFill
    title = f"{country} — Party Overview  ({date.today()})"
    ws.append([title])
    ws[1][0].font = Font(bold=True, size=13)
    ws.append([])

    headers = ["Party", "🟢 Active", "⚫ Inactive", "❌ Removed", "Total Ads",
               "Total Impressions", "Total Spend (€)"]
    write_header(ws, headers, fill, hdr_font)
    ws.row_dimensions[3].height = 22  # header is row 3

    totals = [0] * 6
    for i, (party, active, inactive, removed, total, imp, spend) in enumerate(rows, 1):
        ws.append([party, active, inactive, removed, total, imp, spend])
        row_idx = i + 3
        if alt_fill[i % 2]:
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = alt_fill[i % 2]
        for j, v in enumerate([active, inactive, removed, total, imp, spend]):
            totals[j] += (v or 0)

    # Totals row
    ws.append(["TOTAL"] + totals)
    tr = ws.max_row
    for col in range(1, 8):
        ws.cell(row=tr, column=col).font = total_font
        ws.cell(row=tr, column=col).fill = PatternFill("solid", fgColor="D9E1F2")

    autofit(ws)


def build_top_sheet(ws, rows, title_str, sort_col, fill, hdr_font, alt_fill, link_font):
    from openpyxl.styles import Font, Alignment
    ws.append([title_str])
    ws[1][0].font = Font(bold=True, size=13)
    ws.append([])

    headers = ["Politician", "Party", "Page Name",
               "🟢 Active", "⚫ Inactive", "❌ Removed",
               "Max Spend (€)", "Max Impressions", "Total Spend (€)", "Total Impressions"]
    write_header(ws, headers, fill, hdr_font)

    for i, (query, party, page, active, inactive, removed_cnt,
            max_spend, max_imp, currency, total_spend, total_imp) in enumerate(rows, 1):
        ws.append([fmt_name(query), party, page,
                   active, inactive, removed_cnt,
                   max_spend, max_imp, total_spend, total_imp])
        row_idx = i + 3
        if alt_fill[i % 2]:
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).fill = alt_fill[i % 2]

    autofit(ws)


def build_ads_sheet(ws, rows, title_str, fill, hdr_font, alt_fill, link_font, include_removed_col=False):
    from openpyxl.styles import Font
    ws.append([title_str])
    ws[1][0].font = Font(bold=True, size=13)
    ws.append([])

    if include_removed_col:
        headers = ["Politician", "Party", "District", "Page Name",
                   "Ad ID", "View", "Started", "Removed At",
                   "Impressions", "Spend"]
    else:
        headers = ["Politician", "Party", "District", "Page Name",
                   "Ad ID", "View", "Started",
                   "Impressions", "Spend"]
    write_header(ws, headers, fill, hdr_font)

    for i, row in enumerate(rows, 1):
        ad_id = row[0]
        query = row[1]
        party = row[2]
        dist  = row[3]
        page  = row[4]
        start = row[5]

        if include_removed_col:
            rem_at = (row[6] or "")[:10]
            imp_lo, imp_hi = row[7], row[8]
            sp_lo, sp_hi, cur = row[9], row[10], row[11]
            data = [fmt_name(query), party, dist, page,
                    ad_id, "View", start, rem_at,
                    fmt_range(imp_lo, imp_hi),
                    fmt_range(sp_lo, sp_hi, cur)]
        else:
            imp_lo, imp_hi = row[6], row[7]
            sp_lo, sp_hi, cur = row[8], row[9], row[10]
            data = [fmt_name(query), party, dist, page,
                    ad_id, "View", start,
                    fmt_range(imp_lo, imp_hi),
                    fmt_range(sp_lo, sp_hi, cur)]

        ws.append(data)
        row_idx = i + 3
        # Hyperlink on "View"
        view_col = 6
        c = ws.cell(row=row_idx, column=view_col)
        c.hyperlink = AD_URL.format(ad_id)
        c.font = link_font
        if alt_fill[i % 2]:
            for col in range(1, len(headers) + 1):
                if col != view_col:
                    ws.cell(row=row_idx, column=col).fill = alt_fill[i % 2]

    autofit(ws)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default=BASE, help="Output folder")
    args = parser.parse_args()

    # ── Migrate both DBs upfront — ensures election_related column exists ───
    for db_path in [DB_CY, DB_MT]:
        if os.path.exists(db_path):
            conn = sqlite3.connect(db_path)
            migrate_db(conn)
            conn.close()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        sys.exit("ERROR: pip install openpyxl")

    today     = str(date.today())
    fname     = os.path.join(args.out, f"summary_{today}.xlsx")
    wb        = openpyxl.Workbook()

    # ── colour palette ──────────────────────────────────────────────────────
    FILLS = {
        "blue":   PatternFill("solid", fgColor="1F4E79"),
        "green":  PatternFill("solid", fgColor="375623"),
        "red":    PatternFill("solid", fgColor="A50021"),
        "purple": PatternFill("solid", fgColor="4B0082"),
        "teal":   PatternFill("solid", fgColor="005F5F"),
        "orange": PatternFill("solid", fgColor="8B4500"),
    }
    HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
    ALT_CY     = [None, PatternFill("solid", fgColor="EEF4FF")]
    ALT_MT     = [None, PatternFill("solid", fgColor="FFF8EE")]
    LINK_FONT  = Font(color="0563C1", underline="single", size=9)
    TOTAL_FONT = Font(bold=True, size=10)

    # ── Cyprus ──────────────────────────────────────────────────────────────
    conn_cy = sqlite3.connect(DB_CY, timeout=30)
    conn_cy.execute("PRAGMA journal_mode=WAL")
    try:
        # Sheet 1: CY Party Overview
        ws = wb.active
        ws.title = "CY — Party Overview"
        build_party_sheet(ws, party_overview(conn_cy, today),
                          "🇨🇾 Cyprus", FILLS["blue"], HDR_FONT, ALT_CY, LINK_FONT, TOTAL_FONT)

        # Sheet 2: CY Top by Spend
        ws2 = wb.create_sheet("CY — Top by Spend")
        build_top_sheet(ws2, top_politicians(conn_cy, today, "total_spend"),
                        f"🇨🇾 Cyprus — Top 30 Advertisers by Spend  ({today})",
                        "spend", FILLS["green"], HDR_FONT, ALT_CY, LINK_FONT)

        # Sheet 3: CY Top by Impressions
        ws3 = wb.create_sheet("CY — Top by Impressions")
        build_top_sheet(ws3, top_politicians(conn_cy, today, "total_imp"),
                        f"🇨🇾 Cyprus — Top 30 Advertisers by Impressions  ({today})",
                        "imp", FILLS["teal"], HDR_FONT, ALT_CY, LINK_FONT)

        # Sheet 4: CY Active Ads
        ws4 = wb.create_sheet("CY — Active Ads")
        build_ads_sheet(ws4, active_ads(conn_cy, today),
                        f"🇨🇾 Cyprus — Currently Active Election Ads  ({today})",
                        FILLS["purple"], HDR_FONT, ALT_CY, LINK_FONT)

        # Sheet 5: CY Removed Ads
        ws5 = wb.create_sheet("CY — Removed Ads")
        build_ads_sheet(ws5, removed_ads(conn_cy),
                        f"🇨🇾 Cyprus — Removed Election Ads  (partial — removal check ongoing)",
                        FILLS["red"], HDR_FONT, ALT_CY, LINK_FONT, include_removed_col=True)
    finally:
        conn_cy.close()

    # ── Malta ────────────────────────────────────────────────────────────────
    if os.path.exists(DB_MT):
        conn_mt = sqlite3.connect(DB_MT, timeout=30)
        conn_mt.execute("PRAGMA journal_mode=WAL")
        migrate_db(conn_mt)
        try:
            ws6 = wb.create_sheet("MT — Party Overview")
            build_party_sheet(ws6, party_overview(conn_mt, today),
                              "🇲🇹 Malta", FILLS["orange"], HDR_FONT, ALT_MT, LINK_FONT, TOTAL_FONT)

            ws7 = wb.create_sheet("MT — Active Ads")
            build_ads_sheet(ws7, active_ads(conn_mt, today),
                            f"🇲🇹 Malta — Currently Active Election Ads  ({today})",
                            FILLS["orange"], HDR_FONT, ALT_MT, LINK_FONT)
        finally:
            conn_mt.close()

    wb.save(fname)
    print(f"Saved → {fname}")

    # Print quick console summary
    print(f"\n{'─'*55}")
    print(f"  Generated: summary_{today}.xlsx")
    print(f"  Sheets   : {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"    • {s:<30} {ws.max_row - 3:>5} rows")
    print(f"{'─'*55}")


if __name__ == "__main__":
    main()
