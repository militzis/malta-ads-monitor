"""Export every distinct TikTok account we've found, grouped by category.
   Shows the profile URL so you can click through and verify each."""
import sys, sqlite3, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8')

DB  = r'C:\Users\milit\meta_pipeline_data\politician_ads.db'
OUT = r"C:\Users\milit\Downloads\tiktok_profiles.xlsx"

c = sqlite3.connect(DB)
rows = list(c.execute("""
    SELECT advertiser_id, advertiser_disclosed_name AS handle,
           matched_candidate, matched_party, matched_district,
           COUNT(*) AS ads,
           SUM(CASE WHEN transcript IS NOT NULL AND length(transcript)>20 THEN 1 ELSE 0 END) AS transcribed,
           MIN(first_shown), MAX(last_shown),
           match_type
    FROM tiktok_ads
    GROUP BY advertiser_id, handle
    ORDER BY ads DESC
"""))
print(f"Distinct advertisers: {len(rows)}")

# Annotations from earlier triage (manually verified)
ANNOTATIONS = {
    # CONFIRMED real candidates (transcripts verified political content)
    '7578569963879792657': ('Candidate', 'Παζάρος Χαράλαμπος', 'ΔΗΣΥ', 'Πάφος'),
    '7488289249494679569': ('Candidate', 'Ιωάννου Κλεονίκη', 'ΑΜΔΗ', 'Λεμεσός'),
    '7563644604046852097': ('Candidate', 'Χρυσάνθου Λοΐζος', 'ΑΜΔΗ', 'Λευκωσία'),
    '7533251756613189648': ('Candidate', 'Φλουρέντζου Μάριος', 'ΕΛΑΜ', 'Αμμόχωστος'),
    '7481250791723008017': ('Candidate', 'Ηλιά Μάριος', 'ΔΗΣΥ', 'Λευκωσία'),
    '7554065792900448257': ('Candidate', 'Καλοπαίδης Μιχάλης', 'ΒΟΛΤ', 'Λάρνακα'),
    '7612669198808039440': ('Candidate', 'Πουλλικκάς Μάριος', 'ΕΛΑΜ', 'Λευκωσία'),
    # Concat-match confirmed candidates
    'argentoulaioannou':   ('Candidate', 'Ιωάννου Αργεντούλα', 'ΑΚΕΛ', 'Λεμεσός'),
    'argyrosevangelou':    ('Candidate', 'Ευαγγέλου Αργυρός', 'ΔΗΣΥ', 'Λάρνακα'),
    'chrysanthossavvidis': ('Candidate', 'Σαββίδης Χρύσανθος', 'ΔΗΚΟ', 'Πάφος'),
    'kyprianouanna':       ('Candidate', 'Κυπριανού Άννα', 'ΔΗΠΑ', 'Λευκωσία'),
    'kyproskyprianou4':    ('Candidate', 'Κυπριανού Κύπρος', 'ΑΛΜΑ', 'Λευκωσία'),
    'lakiskonstantinou1':  ('Candidate', 'Κωνσταντίνου Λάκης', 'ΕΔΕΚ', 'Λεμεσός'),
    'pamposkiskonstantinos':('Candidate','Παμπόσκης Κωνσταντίνος', 'ΕΛΑΜ', 'Λεμεσός'),
    'petros.minas2':       ('Candidate', 'Μηνάς Πέτρος', 'ΔΕΚ', 'Λεμεσός'),
    'petrouiakovos':       ('Candidate', 'Πέτρου Ιάκωβος', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ', 'Λευκωσία'),
    'ploutarchosparris':   ('Candidate', 'Παρρής Πλούταρχος', 'ΣΗΚΟΥ ΠΑΝΩ', 'Λευκωσία'),
    'theodosisavgousti':   ('Candidate', 'Αυγουστή Θεοδόσης', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ', 'Λευκωσία'),
    'antonis.antoniou_':   ('Candidate', 'Αντωνή Αντώνης', 'ΕΔΕΚ', 'Λάρνακα'),
    # Post-resolution NEW candidates (found by concat-match on resolved handles)
    'marios_stavrou_':     ('Candidate', 'Σταύρου Μάριος',     'ΑΜΔΗ',         'Λευκωσία'),
    'parismarkou':         ('Candidate', 'Μάρκου Πάρις',       'ΔΗΣΥ',         'Αμμόχωστος'),
    'michalis_fellas':     ('Candidate', 'Φελλάς Μιχάλης',     'ΔΗΣΥ',         'Λεμεσός'),
    'marios.neofytou':     ('Candidate', 'Νεοφύτου Μάριος',    'ΑΛΜΑ',         'Πάφος'),
    'elenachristou1':      ('Candidate', 'Χρίστου Έλεν',       'ΣΗΚΟΥ ΠΑΝΩ',   'Λεμεσός'),
    # User-verified candidates from manual profile review (2026-05-17)
    'deme2023':            ('Candidate', 'Χατζησταύρου Δήμητρα', 'ΑΜΔΗ',         'Λεμεσός'),
    'theanicolaou6':       ('Candidate', 'Νικολάου Θέα',          'ΔΗΠΑ',         'Λευκωσία'),
    'evgenioshamboullas':  ('Candidate', 'Χαμπούλλας Ευγένιος',   'ΕΛΑΜ',         'Λεμεσός'),
    'paliosn':             ('Candidate', 'Πάλιος Νίκος',           'ΑΜΔΗ',         'Πάφος'),
    'elena.vrahimi':       ('Candidate', 'Βραχίμη Έλενα',          'ΔΗΠΑ',         'Λευκωσία'),
    'apostolouaa':         ('Needs verification', 'Αποστόλου (homonym ΛΑΕ/ΔΗΚΟ)', 'ΛΑΕ_or_ΔΗΚΟ', '?'),
    'phivos.doukanaris':   ('Candidate', 'Δουκανάρης Φοίβος',      'ΑΜΔΗ',         'Λάρνακα'),
    'steliosmohicanstylianou': ('Candidate', 'Στυλιανού Στέλιος (Κασιουλή)', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ', 'Πάφος'),
    'apostolouaa':           ('Candidate', 'Αποστόλου Ανδρέας',     'ΔΗΚΟ',         'Λάρνακα'),
    'maria.loizou85':        ('Candidate', 'Λοΐζου Μαρία',           'ΑΜΔΗ',         'Λεμεσός'),
    'adrianachristodoulou13':('Candidate', 'Χριστοδούλου Αδριάνα',   'ΕΛΑΜ',         'Λάρνακα'),
    'lysandrides':           ('Candidate', 'Λυσανδρίδης Γεώργιος',   'ΔΗΣΥ',         'Αμμόχωστος'),
    'steliosstylianou78':    ('Candidate', 'Στυλιανού Στέλιος',       'ΕΛΑΜ',         'Λευκωσία'),
    'antartis':              ('Party supporter', 'ΑΝΤΑΡΤΗΣ — supporter account', 'ΣΗΚΟΥ ΠΑΝΩ', '-'),
    'theofilos891':          ('Party supporter', 'Θεόφιλος — ΕΛΑΜ supporter (torch-march creative)', 'ΕΛΑΜ', '-'),
    'tsitsis_channel':       ('Candidate', 'Τσίτσης Γιάννης',           'ΑΜΔΗ',         'Λεμεσός'),
    'fotinitsiridou':        ('Candidate', 'Τσιρίδου Φωτεινή',          'ΔΗΣΥ',         'Λεμεσός'),
    # Handle-pattern discovery (2026-05-18) — confirmed via bio scan + creative extraction
    'neolaia.lakedaimonioi': ('Party account', 'Πατριωτικό Μέτωπο Λακεδαιμόνιοι — youth wing', 'ΛΑΚΕΔΑΙΜΟΝΙΟΙ', '-'),
    'michalis.kounouni':     ('Candidate', 'Κουνούνης Μιχάλης', 'ΔΗΣΥ', 'Λεμεσός'),
    'fact.check.cyprus':     ('News outlet', 'Fact Check Cyprus — anti-disinformation NGO', '-', '-'),
    'pantelis.vladimirou':   ('Likely false positive', 'webarts.agency co-founder (marketing)', '-', '-'),
    'isaak.solomou':         ('Likely false positive', 'Religious-quote personal account', '-', '-'),
    'charis_av':             ('Party supporter', 'Χάρης — ΣΗΚΟΥ ΠΑΝΩ supporter', 'ΣΗΚΟΥ ΠΑΝΩ', '-'),
    'blackdreamscottage':    ('Party supporter', 'ΑΜΔΗ supporter account', 'ΑΜΔΗ', '-'),
    'phedonphedonos':        ('Politician (non-candidate)', 'Φαίδων Φαίδωνος — Δήμαρχος Πάφου', 'ΔΗΣΥ', 'Πάφος'),
    'adiafthoroi':           ('Commentator', 'Αδιάφθοροι — anti-small-party commentary account', '-', '-'),
    'neadynami':             ('Political movement', 'Νέα Δυναμική — political initiative', 'Νέα Δυναμική', '-'),
    'elena_stefani19':       ('Party supporter', 'Έλενα Στεφανή — ΕΛΑΜ supporter', 'ΕΛΑΜ', '-'),
    'yiannis.laouris':       ('Needs verification', 'Γιάννης Λαουρής — candidacy announced, party TBD (ΣΗΚΟΥ ΠΑΝΩ likely from rhetoric)', '?', '?'),
    'andreamanoli7':         ('Candidate', 'Θεολόγου Μανωλή Άνδρεα', 'ΔΗΣΥ', 'Λάρνακα'),
    'peripaizoun.sas':       ('Likely false positive', 'Irrelevant (user-flagged)', '-', '-'),
    'triplesevenfc':         ('Likely false positive', 'Irrelevant (user-flagged)', '-', '-'),
    'djpietro_official':     ('Likely false positive', 'DJ Pietro — Cypriot folk-music release (κυπριακό τραγούδι, not Cyprus problem)', '-', '-'),
    # Batch resolution 2026-05-18 — handles resolved from numeric BIDs
    'kyriakimanousaki':      ('Candidate', 'Μανουσάκη Κυριακή', 'ΕΔΕΚ', 'Λευκωσία'),
    'veronikapavlidou':      ('Candidate', 'Παυλίδου Βερόνικα', 'ΔΕΚ', 'Αμμόχωστος'),
    'orfeas.restaurant':     ('Likely false positive', 'Orfeas restaurant', '-', '-'),
    'faithtofaithdaily':     ('Likely false positive', 'Religious devotional account', '-', '-'),
    'mini.armoire':          ('Likely false positive', 'Clothing store', '-', '-'),
    'cacoyannis':            ('Likely false positive', 'Cacoyannis Foundation — culture/cinema', '-', '-'),
    'habibiyialla.a3':       ('Likely false positive', 'Arabic music account', '-', '-'),
    'ortiz.by.mariposa':     ('Likely false positive', 'Fashion brand', '-', '-'),
    'ikonestispisteos':      ('Likely false positive', 'Orthodox-icons retailer', '-', '-'),
    'tarintensofraya':       ('Likely false positive', 'Personal/unknown account', '-', '-'),
    'euroresidence.cy':      ('Likely false positive', 'Real estate', '-', '-'),
    'djzeph2k':              ('Likely false positive', 'DJ Zeph', '-', '-'),
    'cricospr':              ('Likely false positive', 'Personal/unknown account', '-', '-'),
    # Second batch 2026-05-18
    'voulakokkinou7':        ('Candidate', 'Κοκκίνου Παρασκευή ("Βούλα")', 'ΑΛΜΑ', 'Λευκωσία'),
    'proeklogikipanagioti':  ('Needs verification', 'Προεκλογική Παναγιώτη — 13 candidates named Π/Παναγιώτης, needs bio scan', '?', '?'),
    'synotruck':             ('Likely false positive', 'Truck dealer', '-', '-'),
    'tek.event':             ('Likely false positive', 'Event production company', '-', '-'),
    'efebozgeyikk':          ('Likely false positive', 'Turkish-Cypriot name (north CY)', '-', '-'),
    'monde.gr':              ('Likely false positive', 'Greek news outlet (Greece, not CY)', '-', '-'),
    'eccekocc':              ('Likely false positive', 'Turkish-sounding personal', '-', '-'),
    'jimboscryptoland':      ('Likely false positive', 'Crypto account', '-', '-'),
    'charalambos_michael':   ('Likely false positive', 'Musician — song titled "Αμμόχωστος"', '-', '-'),
    'chrisdemetriou26':      ('Likely false positive', 'Personal vlogger (year suffix is coincidence)', '-', '-'),
    's.pavlides':            ('Likely false positive', 'Bioiatriki Lab — wellness brand', '-', '-'),
    'reset_cy':              ('Likely false positive', 'Social-issues NGO', '-', '-'),
    'horizon.cyparty':       ('Likely false positive', 'Summer 2026 mountain event', '-', '-'),
    'chrysalexofficial':     ('Likely false positive', 'Tech entrepreneur / public speaker', '-', '-'),
    'marilena_antoniadou':   ('Likely false positive', 'Personal Instagram crosspost', '-', '-'),
    # Discovery-sweep false positives (2026-05-18) — confirmed via bio scan
    'kadisestates':          ('Likely false positive', 'Kadis Estates — real estate (kadis.com.cy), homonym surname', '-', '-'),
    'nataly.laser.house':    ('Likely false positive', 'Candela Alexandrite laser clinic, Strovolos', '-', '-'),
    'demetris.aletraris':    ('Candidate', 'Αλετράρης Δημήτρης',     'ΑΜΔΗ',         'Λευκωσία'),
    # Bio-scan auto-discovered candidates (2026-05-18)
    'allasoneagian':         ('Candidate', 'Παπαχριστοφόρου Στυλιανός', 'ΑΜΔΗ',     'Αμμόχωστος'),
    'andreas_papacharalambous':('Candidate','Παπαχαραλάμπους Ανδρέας', 'ΕΛΑΜ',      'Λευκωσία'),
    'florentzos_karayiannas':('Candidate', 'Καραγιάννας Φλωρέντζος',   'ΔΗΚΟ',     'Λεμεσός'),
    'demosg_cy':             ('Candidate', 'Γεωργιάδης Δήμος',          'ΔΗΣΥ',     'Κερύνεια'),
    'stamos.papavasili':     ('Candidate', 'Παπαβασιλείου Σταμάτης',    'ΟΙΚΟΛΟΓΟΙ','Λεμεσός'),
    'modecristo':            ('Candidate', 'Μοδέστου Χρίστος',          'ΟΙΚΟΛΟΓΟΙ','Λεμεσός'),
    'tasos2026':             ('Candidate', 'Αναστασίου Αναστάσιος',     'ΑΜΔΗ',     'Αμμόχωστος'),
    'paraschou1':            ('Candidate', 'Παρασχού Αντώνης',          'ΣΗΚΟΥ ΠΑΝΩ', 'Αμμόχωστος'),
    'ari.habeshian':         ('Candidate', 'Χαπεσιάν Άρι',               'ΔΗΠΑ',     'Λευκωσία'),
    'pooldoctorcyprus':      ('Likely false positive', 'Pool maintenance business', '-', '-'),
    'pepsis_and_son':        ('Likely false positive', 'Construction business', '-', '-'),
    'rentspotcy':          ('Likely false positive', 'Rental business', '-', '-'),
    # Homonym advertisers — need profile verification
    'steliosmohicanstylianou': ('Candidate', 'Στυλιανού Στέλιος ("O Souvlakis")', 'ΕΝΕΡΓΟΙ ΠΟΛΙΤΕΣ', 'Πάφος'),
    'steliosstylianou78':      ('Candidate', 'Στυλιανού Στέλιος', 'ΕΛΑΜ', 'Λευκωσία'),
    'maria.loizou85':          ('Candidate', 'Λοΐζου Μαρία', 'ΑΜΔΗ', 'Λεμεσός'),
    # Confirmed FP
    'antonis.cleaning':    ('Likely false positive', 'Cleaning business (homonymous first name)', '-', '-'),
    'mariannaathanasiou':  ('Candidate', 'Αθανασίου Μαριάννα', 'ΛΑΚΕΔΑΙΜΟΝΙΟΙ', 'Λεμεσός'),
    'mariapastella1':      ('Candidate', 'Παστέλλα Μαρία', 'ΕΛΑΜ', 'Λεμεσός'),
    'stellapetridou5':     ('Candidate', 'Πετρίδου Στέλλα', 'ΟΙΚΟΛΟΓΟΙ', 'Λευκωσία'),
    # Party-level + non-candidate political accounts (keyed by bid since DB has numeric handles)
    '7628713809988960272': ('Party account', '@almalimassol — ΑΛΜΑ regional', 'ΑΛΜΑ', 'Λεμεσός'),
    '7450088911596011536': ('Party coordinator', '@nakiskyriakou — 44 ads of ΣΗΚΟΥ ΠΑΝΩ slogan', 'ΣΗΚΟΥ ΠΑΝΩ', '-'),
    '7600116160427851793': ('News outlet', '@tanea.cy — Ta Nea CY', '-', '-'),
    '7464520683175804929': ('Podcast',     '@tolkerscy — Talkers CY podcast', '-', '-'),
    '7444996016945774609': ('Commentator', '@gnosths_ths_istorias — history/politics commentary', '-', '-'),
    '7479401042212487184': ('Satirist',    '@gastrimargos — food + political satire', '-', '-'),
    '7454901333884141584': ('Unverified',  '@angelosiacovides — needs profile check', '?', '?'),
    # Tier A confirmed candidates (concat-match found via Playwright resolution)
    '7603797428927578113': ('Candidate', '@mariapastella1 — Παστέλλα Μαρία', 'ΕΛΑΜ', 'Λεμεσός'),
    '7604116497765711889': ('Candidate', '@stellapetridou5 — Πετρίδου Στέλλα', 'ΟΙΚΟΛΟΓΟΙ', 'Λευκωσία'),
    '7579460984457150480': ('Candidate', '@mariannaathanasiou — Αθανασίου Μαριάννα', 'ΛΑΚΕΔΑΙΜΟΝΙΟΙ', 'Λεμεσός'),
}

def categorize(bid, handle, ads_count, match_type):
    bid = str(bid)
    h = (handle or '').lower()
    if bid in ANNOTATIONS:    return ANNOTATIONS[bid]
    if h in ANNOTATIONS:      return ANNOTATIONS[h]
    if match_type == 'manual_resume':
        return ('Candidate (unverified)', '?', '?', '?')
    if match_type == 'needs_profile_verification':
        return ('Needs verification', '?', '?', '?')
    if match_type and match_type.startswith('content_keyword_tier_A'):
        return ('Tier A (political-content)', '?', '?', '?')
    if match_type == 'content_keyword' or (match_type and 'content_keyword' in match_type):
        return ('Content-keyword hit', '?', '?', '?')
    if match_type and 'false_positive' in match_type:
        return ('Likely false positive', '-', '-', '-')
    return ('?', '?', '?', '?')

# Build workbook
wb = Workbook()
ws = wb.active
ws.title = 'Profiles'

HEADERS = ['Category', 'Handle', 'Profile URL', 'Identified as', 'Party', 'District',
           '# ads', '# with transcript', 'First shown', 'Last shown', 'match_type']
HEADER_FILL = PatternFill('solid', fgColor='305496')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
HYPER_FONT  = Font(color='0563C1', underline='single')
FILLS = {
    'Candidate':              PatternFill('solid', fgColor='C6E0B4'),
    'Candidate (unverified)': PatternFill('solid', fgColor='E2EFDA'),
    'Party account':          PatternFill('solid', fgColor='BDD7EE'),
    'Party coordinator':      PatternFill('solid', fgColor='BDD7EE'),
    'Party supporter':        PatternFill('solid', fgColor='D9E1F2'),
    'News outlet':            PatternFill('solid', fgColor='FFF2CC'),
    'Podcast':                PatternFill('solid', fgColor='FFF2CC'),
    'Commentator':            PatternFill('solid', fgColor='FFF2CC'),
    'Satirist':               PatternFill('solid', fgColor='FFF2CC'),
    'Needs verification':     PatternFill('solid', fgColor='FCE4D6'),
    'Tier A (political-content)': PatternFill('solid', fgColor='FCE4D6'),
    'Unverified':             PatternFill('solid', fgColor='FCE4D6'),
    'Content-keyword hit':    PatternFill('solid', fgColor='FFFFFF'),
    'Likely false positive':  PatternFill('solid', fgColor='F4B0B0'),
    '?':                      PatternFill('solid', fgColor='FFFFFF'),
}
CATEGORY_SORT = ['Candidate', 'Candidate (unverified)', 'Party account', 'Party coordinator',
                 'Party supporter',
                 'News outlet', 'Podcast', 'Commentator', 'Satirist',
                 'Needs verification', 'Tier A (political-content)', 'Unverified',
                 'Content-keyword hit', 'Likely false positive', '?']

# Header
for ci, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT

# Sort by category then ads
def sort_key(r):
    bid, handle, _, _, _, ads, _, _, _, mt = r
    cat, _, _, _ = categorize(bid, handle, ads, mt)
    try: idx = CATEGORY_SORT.index(cat)
    except ValueError: idx = 999
    return (idx, -ads)
rows_sorted = sorted(rows, key=sort_key)

r_idx = 2
counts = defaultdict(int)
for bid, handle, cand, party, district, ads, trans, first, last, mt in rows_sorted:
    cat, name, p, dist = categorize(bid, handle, ads, mt)
    counts[cat] += 1
    profile_url = (f"https://www.tiktok.com/@{handle}"
                   if handle and not str(handle).isdigit() else "")
    values = [
        cat, handle, profile_url,
        name if name not in ('?','-') else (cand or ''),
        p if p not in ('?','-') else (party or ''),
        dist if dist not in ('?','-') else (district or ''),
        ads, trans, first or '', last or '', mt or '',
    ]
    fill = FILLS.get(cat)
    for ci, v in enumerate(values, 1):
        cell = ws.cell(row=r_idx, column=ci, value=v)
        if fill: cell.fill = fill
        cell.alignment = Alignment(vertical='top', wrap_text=False)
    # Make profile URL clickable
    if profile_url:
        pc = ws.cell(row=r_idx, column=3)
        pc.hyperlink = profile_url
        pc.font = HYPER_FONT
    r_idx += 1

# Column widths
WIDTHS = {'Category':28,'Handle':28,'Profile URL':40,'Identified as':32,
          'Party':18,'District':14,'# ads':6,'# with transcript':10,
          'First shown':11,'Last shown':11,'match_type':36}
for i, h in enumerate(HEADERS, 1):
    ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(h, 16)
ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions

# Summary tab
ws2 = wb.create_sheet('Summary')
ws2.append(['Category', '# advertisers'])
for cat in CATEGORY_SORT:
    if counts.get(cat):
        ws2.append([cat, counts[cat]])
for cell in ws2[1]:
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 15

try:
    wb.save(OUT)
    print(f"Saved {OUT}")
except PermissionError:
    from datetime import datetime
    alt = OUT.replace('.xlsx', f"_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(alt); print(f"[locked] Saved {alt}")

print("\nBreakdown:")
for cat in CATEGORY_SORT:
    if counts.get(cat):
        print(f"  {counts[cat]:>4}  {cat}")
